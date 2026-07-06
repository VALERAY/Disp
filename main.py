import sqlite3
import re
import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
from datetime import datetime, timedelta
from tkcalendar import DateEntry
from tkinter import filedialog
import ctypes
import glob
import os
import sys
from pathlib import Path
import zipfile
import xml.etree.ElementTree as ET

# Опциональный импорт pandas (используется только для экспорта)
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

def get_month_year_label(db_filename):
    import re
    m = re.match(r'app_(\d{4})_(\d{2})\.db', db_filename)
    if not m:
        return db_filename
    year = int(m.group(1))
    month = int(m.group(2))
    months_ru = [
        "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
    ]
    return f"{months_ru[month]} {year}"

# ===== Создание базы =====
# ...existing code...

def _get_app_dir() -> Path:
    """Папка приложения. Для .exe — папка, где лежит exe; для .py — текущая папка."""
    try:
        if getattr(sys, 'frozen', False):  # PyInstaller
            return Path(sys.executable).parent
    except Exception:
        pass
    return Path.cwd()


def _load_db_base_dir() -> str:
    """Возвращает каталог хранения помесячных БД. Хранится в .db_dir (иначе текущая папка)."""
    try:
        cfg = _get_app_dir() / ".db_dir"
        if cfg.exists():
            p = cfg.read_text(encoding="utf-8").strip().strip('"')
            if p:
                return str(Path(p).expanduser().resolve())
    except Exception:
        pass
    return str(_get_app_dir())


def _save_db_base_dir(directory: str) -> None:
    try:
        (_get_app_dir() / ".db_dir").write_text(str(directory), encoding="utf-8")
    except Exception:
        pass


def get_db_name_by_date(date_str=None):
    if not date_str:
        now = datetime.now()
        year, month = now.year, now.month
    else:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        year, month = dt.year, dt.month
    return f"app_{year}_{month:02d}.db"


def get_db_full_path_by_date(date_str=None) -> str:
    base_dir = _load_db_base_dir()
    return str(Path(base_dir) / get_db_name_by_date(date_str))

current_db_file = get_db_full_path_by_date()
Path(current_db_file).parent.mkdir(parents=True, exist_ok=True)
conn = sqlite3.connect(current_db_file)
cursor = conn.cursor()

def _get_common_db_path() -> str:
    """Полный путь к постоянной БД пользователей (в том же каталоге, что и помесячные)."""
    base_dir = _load_db_base_dir()
    return str(Path(base_dir) / "app.db")

# Присоединяем постоянную базу для пользователей из выбранного каталога (поддерживает сетевые пути)
try:
    _common_path = Path(_get_common_db_path())
    _common_path.parent.mkdir(parents=True, exist_ok=True)
    # Экранируем одинарные кавычки для безопасной подстановки в SQL
    _common_escaped = str(_common_path).replace("'", "''")
    conn.execute(f"ATTACH DATABASE '{_common_escaped}' AS common")
except Exception:
    pass

def switch_db(db_file):
    global conn, cursor, current_db_file
    if conn:
        conn.close()
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    current_db_file = db_file
    # Повторно присоединяем постоянную базу для пользователей после переключения
    try:
        _common_path = Path(_get_common_db_path())
        _common_path.parent.mkdir(parents=True, exist_ok=True)
        _common_escaped = str(_common_path).replace("'", "''")
        conn.execute(f"ATTACH DATABASE '{_common_escaped}' AS common")
    except Exception:
        pass
    ensure_schema()
# ...existing code...

cursor.execute("""CREATE TABLE IF NOT EXISTS common.users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT UNIQUE,
    password TEXT,
    role TEXT
)""")
cursor.execute("""CREATE TABLE IF NOT EXISTS records (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT,
    surname TEXT,
    description TEXT,
    date TEXT,
    user_id INTEGER
)""")
conn.commit()

# ===== Миграция схемы =====
def ensure_schema_for_connection(connection, cursor_to_use):
    """Обновляет схему базы данных для указанного подключения."""
    cursor_to_use.execute("PRAGMA table_info(records)")
    existing_columns = [row[1] for row in cursor_to_use.fetchall()]
    if "assignment_date" not in existing_columns:
        cursor_to_use.execute("ALTER TABLE records ADD COLUMN assignment_date TEXT")
    if "status" not in existing_columns:
        cursor_to_use.execute("ALTER TABLE records ADD COLUMN status TEXT")
        cursor_to_use.execute("UPDATE records SET status='не выполнено' WHERE status IS NULL")
    else:
        try:
            cursor_to_use.execute("UPDATE records SET status='не выполнено' WHERE status='не начато'")
        except Exception:
            pass
    # New fields for problem, phone, address
    if "problem" not in existing_columns:
        cursor_to_use.execute("ALTER TABLE records ADD COLUMN problem TEXT")
    if "phone" not in existing_columns:
        cursor_to_use.execute("ALTER TABLE records ADD COLUMN phone TEXT")
    if "address" not in existing_columns:
        cursor_to_use.execute("ALTER TABLE records ADD COLUMN address TEXT")
    if "created_at" not in existing_columns:
        cursor_to_use.execute("ALTER TABLE records ADD COLUMN created_at TEXT")
    # Раздел благоустройства
    if "improvement" not in existing_columns:
        cursor_to_use.execute("ALTER TABLE records ADD COLUMN improvement TEXT")
    if "brigade_number" not in existing_columns:
        cursor_to_use.execute("ALTER TABLE records ADD COLUMN brigade_number TEXT")
    if "category" not in existing_columns:
        cursor_to_use.execute("ALTER TABLE records ADD COLUMN category TEXT")

    # ===== 1НФ: разбиваем problem на problem_text + deadline_hours =====
    if "problem_text" not in existing_columns:
        cursor_to_use.execute("ALTER TABLE records ADD COLUMN problem_text TEXT")
    if "deadline_hours" not in existing_columns:
        cursor_to_use.execute("ALTER TABLE records ADD COLUMN deadline_hours INTEGER")
    # Мигрируем в Python — надёжнее чем сложный SQL
    try:
        cursor_to_use.execute(
            "SELECT id, problem FROM records WHERE problem_text IS NULL AND problem IS NOT NULL"
        )
        for row_id, prob in cursor_to_use.fetchall():
            if not prob:
                continue
            m = re.search(r'\(\s*срок\s+выполнения\s+(\d+)\s*ч\s*\)', prob, re.IGNORECASE)
            if m:
                p_text = re.sub(r'\s*\(\s*срок\s+выполнения\s+\d+\s*ч\s*\)', '',
                                prob, flags=re.IGNORECASE).strip()
                d_hours = int(m.group(1))
            else:
                p_text = prob.strip()
                d_hours = None
            cursor_to_use.execute(
                "UPDATE records SET problem_text=?, deadline_hours=? WHERE id=?",
                (p_text, d_hours, row_id)
            )
    except Exception:
        pass

    # ===== 1НФ: разбиваем status на status_clean + status_changed_at =====
    if "status_clean" not in existing_columns:
        cursor_to_use.execute("ALTER TABLE records ADD COLUMN status_clean TEXT")
    if "status_changed_at" not in existing_columns:
        cursor_to_use.execute("ALTER TABLE records ADD COLUMN status_changed_at TEXT")
    # Мигрируем существующие данные: "выполнено (16.04.2026 10:00)" → два поля
    cursor_to_use.execute("""
        UPDATE records SET
            status_clean = CASE
                WHEN status LIKE '%(%)%'
                    THEN TRIM(SUBSTR(status, 1, INSTR(status, '(') - 1))
                ELSE TRIM(COALESCE(status, ''))
            END,
            status_changed_at = CASE
                WHEN status LIKE '%(%)%'
                    THEN TRIM(REPLACE(REPLACE(
                        SUBSTR(status, INSTR(status, '(')), '(', ''), ')', ''))
                ELSE NULL
            END
        WHERE status_clean IS NULL AND status IS NOT NULL
    """)

    # ===== 1НФ: таблица категорий (справочник) =====
    try:
        cursor_to_use.execute("""
            CREATE TABLE IF NOT EXISTS common.categories (
                id   INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL
            )
        """)
        # Заполняем справочник из уже существующих записей
        cursor_to_use.execute("""
            INSERT OR IGNORE INTO common.categories (name)
            SELECT DISTINCT category FROM records
            WHERE category IS NOT NULL AND category != ''
        """)
    except Exception:
        pass

    if "category_id" not in existing_columns:
        cursor_to_use.execute("ALTER TABLE records ADD COLUMN category_id INTEGER")
    # Проставляем category_id для существующих записей
    try:
        cursor_to_use.execute("""
            UPDATE records SET category_id = (
                SELECT id FROM common.categories WHERE name = records.category
            )
            WHERE category_id IS NULL AND category IS NOT NULL AND category != ''
        """)
    except Exception:
        pass

    connection.commit()

def ensure_schema():
    """Обновляет схему для текущего глобального подключения."""
    ensure_schema_for_connection(conn, cursor)

ensure_schema()

# ===== Настройка масштабирования =====
def setup_scaling():
    """Настройка автоматического масштабирования для Windows"""
    try:
        # Получаем DPI монитора
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
        # Альтернативный способ для старых версий Windows
        ctypes.windll.user32.SetProcessDPIAware()
    except:
        pass
    
    # Базовые стили (светлая тема)
    style = ttk.Style()
    try:
        style.theme_use('clam')
    except:
        pass

    primary_bg = '#f5f7fb'
    card_bg = '#ffffff'
    accent = '#1e88e5'
    text_color = '#1f2937'

    style.configure('.', background=primary_bg)
    style.configure('TButton', padding=8, font=('Segoe UI', 10), foreground='white', background=accent)
    style.map('TButton', background=[('active', '#1565c0')])
    style.configure('TLabel', font=('Segoe UI', 10), foreground=text_color, background=primary_bg)
    style.configure('TEntry', padding=6, font=('Segoe UI', 10))
    style.configure('TCombobox', padding=3, font=('Segoe UI', 10))
    style.configure('TLabelframe', font=('Segoe UI', 10, 'bold'), background=primary_bg)
    style.configure('TLabelframe.Label', font=('Segoe UI', 10, 'bold'), background=primary_bg, foreground=text_color)

    # Таблица
    style.configure('Treeview', rowheight=26, font=('Segoe UI', 10))
    style.configure('Treeview.Heading', font=('Segoe UI Semibold', 10))
    # Контрастное выделение строки: тёмно-синий фон и белый текст
    try:
        style.map('Treeview',
                  background=[('selected', '#1565c0')],
                  foreground=[('selected', '#ffffff')])
    except Exception:
        try:
            style.map('Treeview', background=[('selected', '#1565c0')])
        except Exception:
            pass

def fit_and_center_window(window, min_width=None, min_height=None, max_ratio=0.9):
    """Подгоняет окно под содержимое и центрирует его на экране.
    max_ratio ограничивает размер окна долей экрана, чтобы не выходило за края.
    """
    try:
        window.update_idletasks()
        req_w = window.winfo_reqwidth()
        req_h = window.winfo_reqheight()
        if min_width:
            req_w = max(req_w, min_width)
        if min_height:
            req_h = max(req_h, min_height)
        screen_w = window.winfo_screenwidth()
        screen_h = window.winfo_screenheight()
        w = min(req_w + 40, int(screen_w * max_ratio))
        h = min(req_h + 60, int(screen_h * max_ratio))
        x = (screen_w // 2) - (w // 2)
        y = (screen_h // 2) - (h // 2)
        window.geometry(f"{w}x{h}+{x}+{y}")
        window.minsize(w, h)
    except Exception:
        pass

def load_resized_image(path: str, target_height: int = 24):
    """Загружает изображение и масштабирует до нужной высоты.
    Возвращает объект PhotoImage (ImageTk.PhotoImage или tk.PhotoImage) либо None.
    """
    try:
        from PIL import Image, ImageTk
        img = Image.open(path)
        if target_height > 0 and img.height != target_height:
            ratio = target_height / float(img.height)
            new_size = (max(1, int(img.width * ratio)), target_height)
            img = img.resize(new_size, Image.LANCZOS)
        return ImageTk.PhotoImage(img)
    except Exception:
        try:
            raw = tk.PhotoImage(file=path)
            if raw.height() > 0 and target_height > 0 and raw.height() > target_height:
                factor = max(1, int(round(raw.height() / target_height)))
                raw = raw.subsample(factor, factor)
            return raw
        except Exception:
            return None

def open_add_user_window():
    def add_new_user():
        uname = entry_new_username.get()
        upass = entry_new_password.get()
        urole = role_var.get()
        if uname and upass:
            try:
                cursor.execute("INSERT INTO common.users (username, password, role) VALUES (?, ?, ?)",
                               (uname, upass, urole))
                conn.commit()
                messagebox.showinfo("Успех", f"Пользователь {uname} добавлен")
                new_user_win.destroy()
            except sqlite3.IntegrityError:
                messagebox.showwarning("Ошибка", "Такой логин уже существует")
        else:
            messagebox.showwarning("Ошибка", "Заполните все поля")

    new_user_win = tk.Toplevel()
    setup_scaling()
    new_user_win.title("Добавить пользователя")
    new_user_win.geometry("300x200")
    new_user_win.resizable(True, True)
    
    # Центрируем окно и подгоняем размеры
    try:
        fit_and_center_window(new_user_win, min_width=300, min_height=200)
    except:
        pass

    main_frame = ttk.Frame(new_user_win, padding="20")
    main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
    
    new_user_win.columnconfigure(0, weight=1)
    new_user_win.rowconfigure(0, weight=1)
    main_frame.columnconfigure(1, weight=1)

    ttk.Label(main_frame, text="Логин:").grid(row=0, column=0, sticky=tk.W, pady=5)
    entry_new_username = ttk.Entry(main_frame, width=25)
    entry_new_username.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

    ttk.Label(main_frame, text="Пароль:").grid(row=1, column=0, sticky=tk.W, pady=5)
    entry_new_password = ttk.Entry(main_frame, show="*", width=25)
    entry_new_password.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

    ttk.Label(main_frame, text="Роль:").grid(row=2, column=0, sticky=tk.W, pady=5)
    role_var = tk.StringVar(value="user")
    role_combo = ttk.Combobox(main_frame, textvariable=role_var, values=["user", "admin"], state="readonly", width=22)
    role_combo.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

    ttk.Button(main_frame, text="Добавить", command=add_new_user).grid(row=3, column=0, columnspan=2, pady=20)


# ===== Добавляем тестового админа =====
cursor.execute("SELECT * FROM common.users WHERE username='admin'")
if not cursor.fetchone():
    cursor.execute("INSERT INTO common.users (username, password, role) VALUES (?, ?, ?)", 
                   ("admin", "admin", "admin"))
    conn.commit()

# ===== Авторизация =====
def login():
    username = entry_login.get()
    password = entry_password.get()
    cursor.execute("SELECT * FROM common.users WHERE username=? AND password= ?", (username, password))
    user = cursor.fetchone()
    if user:
        # Сохраняем логин, если включен чекбокс
        try:
            if remember_var.get():
                with open(".remember_user", "w", encoding="utf-8") as f:
                    f.write((username or "") + "\n" + (password or ""))
            else:
                try:
                    if os.path.exists(".remember_user"):
                        os.remove(".remember_user")
                except Exception:
                    pass
        except Exception:
            pass
        login_window.destroy()
        open_main_window(user)
    else:
        messagebox.showwarning("Ошибка", "Неверный логин или пароль")

# ===== Основное окно =====
def open_main_window(user):
    main = tk.Tk()
    main.title(f"Приложение (пользователь: {user[1]})")
    setup_scaling()
    
    # Автоподгон и центрирование
    main.resizable(True, True)
    # Открываем главное окно сразу в развёрнутом виде (Windows)
    try:
        main.state('zoomed')
    except Exception:
        pass
    try:
        fit_and_center_window(main, min_width=1000, min_height=700)
    except:
        pass
    
    # Настройка весов для масштабирования
    main.columnconfigure(0, weight=1)
    main.rowconfigure(2, weight=1)  # Список записей будет расширяться
    
    # ----   шапка ----
    top_frame = ttk.Frame(main)
    top_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=16, pady=12)
    top_frame.columnconfigure(1, weight=1)

    # Устанавливаем иконку окна как логотип
    try:
        icon_img = load_resized_image("logo2.png", target_height=32)
        if icon_img is not None:
            main.iconphoto(True, icon_img)
            main._icon_img = icon_img  # удерживаем ссылку
    except Exception:
        pass

    # Логотип слева от заголовка (если файл доступен) с аккуратным масштабированием
    logo_img = load_resized_image("logo2.png", target_height=24)
    if logo_img is not None:
        logo_lbl = ttk.Label(top_frame, image=logo_img)
        logo_lbl.image = logo_img  # сохранить ссылку, чтобы изображение не удалилось сборщиком мусора
        logo_lbl.grid(row=0, column=0, sticky="w", padx=(0, 8))

    title_lbl = ttk.Label(top_frame, text="Диспетчер заявок", font=('Segoe UI Semibold', 16))
    title_lbl.grid(row=0, column=1, sticky="w")

    user_lbl = ttk.Label(top_frame, text=f"Пользователь: {user[1]}")
    user_lbl.grid(row=0, column=2, sticky="e", padx=(10, 0))

    if user[3] == "admin":
        ttk.Button(top_frame, text="Добавить пользователя", command=open_add_user_window)\
            .grid(row=0, column=3, sticky="e", padx=(12, 0))

        def open_admin_panel():
            warn = (
                "Вы точно хотите открыть админ-панель?\n\n"
                "Внимание: возможны необратимые последствия при неверных действиях."
            )
            if not messagebox.askyesno("Предупреждение", warn):
                return

            admin_win = tk.Toplevel(main)
            setup_scaling()
            admin_win.title("Админ панель")
            try:
                fit_and_center_window(admin_win, min_width=520, min_height=220)
            except Exception:
                pass

            frm = ttk.LabelFrame(admin_win, text="Настройки базы данных", padding="16")
            frm.grid(row=0, column=0, padx=12, pady=12, sticky=(tk.N, tk.S, tk.E, tk.W))
            admin_win.columnconfigure(0, weight=1)
            admin_win.rowconfigure(0, weight=1)
            frm.columnconfigure(1, weight=1)

            ttk.Label(frm, text="Текущий каталог хранения помесячных БД:").grid(row=0, column=0, sticky=tk.W)
            db_dir_var = tk.StringVar(value=_load_db_base_dir())
            db_dir_entry = ttk.Entry(frm, textvariable=db_dir_var)
            db_dir_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0))

            def choose_dir():
                d = filedialog.askdirectory(title="Выбор каталога для БД")
                if d:
                    db_dir_var.set(d)

            ttk.Button(frm, text="Выбрать...", command=choose_dir).grid(row=0, column=2, padx=(10, 0))

            def save_dir():
                new_dir = db_dir_var.get().strip()
                if not new_dir:
                    messagebox.showwarning("Ошибка", "Укажите каталог")
                    return
                try:
                    Path(new_dir).mkdir(parents=True, exist_ok=True)
                except Exception as e:
                    messagebox.showerror("Ошибка", f"Не удалось создать каталог: {e}")
                    return
                _save_db_base_dir(new_dir)
                messagebox.showinfo("Готово", "Каталог сохранён. Перезапустите выбор базы или приложение.")

            btns = ttk.Frame(admin_win)
            btns.grid(row=1, column=0, padx=12, pady=(0, 12), sticky="e")
            ttk.Button(btns, text="Сохранить", command=save_dir).grid(row=0, column=0, padx=6)
            ttk.Button(btns, text="Закрыть", command=admin_win.destroy).grid(row=0, column=1, padx=6)

        ttk.Button(top_frame, text="Админ панель", command=open_admin_panel)\
            .grid(row=0, column=4, sticky="e", padx=(12, 0))

    # Кнопка выхода на экран входа
    def do_logout():
        try:
            main.destroy()
        finally:
            try:
                open_login_window()
            except Exception:
                pass

    logout_col = 5 if user[3] == "admin" else 3
    ttk.Button(top_frame, text="Выход", command=do_logout).grid(row=0, column=logout_col, sticky="e", padx=(12, 0))


# ---- Добавление записей ----
    frame_add = ttk.LabelFrame(main, text="Добавить запись", padding="12")
    frame_add.grid(row=1, column=0, columnspan=2, padx=16, pady=8, sticky="ew")
    frame_add.columnconfigure(1, weight=1)

    ttk.Label(frame_add, text="Наименование:").grid(row=0, column=0, sticky=tk.W, pady=5)
    name_entry = ttk.Entry(frame_add, width=40)
    name_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

    ttk.Label(frame_add, text="Фамилия:").grid(row=1, column=0, sticky=tk.W, pady=5)
    surname_entry = ttk.Entry(frame_add, width=40)
    surname_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

    # Функция для извлечения данных из Word документа
    def load_problem_categories_from_docx():
        """Загружает problem_categories и problem_to_blank_category из Word документа"""
        doc_path = os.path.join(_get_app_dir(), "Таблица содержание 28.11.2025 ( исправление).docx")
        
        # Значения по умолчанию
        default_categories = {
            "водоотведение": [
                "забой канализационного колодца",
                "течь канализации по дороге",
            ],
            "водоснабжение": [
                "течь воды",
                "течь в / колонки",
                "течь из-под земли",
                "течь пожарного гидранта",
                "течь из -под асфальта",
                "течь трассы холодной воды",
                "откачка воды из колодца",
                "дефект водоразборной колонки",
                "ржавая холодная вода в жилом фонде",
                "Не работает в / колонка",
                "слабое давление х/в",
                "восстановить в/колонку",
                "течь в/к",
                "нет х/в"
            ],
            "общие": [
                "перекладка",
                "водопровод",
                "частные врезки",
                "открыт колодец (отсутствие/несоответствие крышки)",
                "восстановление асфальтобетонного покрытия",
                "разрушена плита колодца",
                "привести к/к в нормативное состояние",
                "привести в / к в нормативное состояние",
                "обвал в/к",
                "обвал к/к",
            ]
        }
        
        default_mapping = {
            "течь канализации по дороге": "канализация",
        }
        
        # Пытаемся загрузить из Word документа
        if os.path.exists(doc_path):
            try:
                content_to_category = {}
                all_contents = []
                
                with zipfile.ZipFile(doc_path, 'r') as docx:
                    xml_content = docx.read('word/document.xml')
                    root = ET.fromstring(xml_content)
                    
                    ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
                    rows = root.findall('.//w:tr', ns)
                    
                    for row in rows:
                        cells = row.findall('.//w:tc', ns)
                        cell_texts = []
                        for cell in cells:
                            texts = [t.text for t in cell.findall('.//w:t', ns) if t.text]
                            cell_texts.append(' '.join(texts).strip())
                        
                        if not any(cell_texts):
                            continue
                        
                        content = ""
                        category = ""
                        
                        if cell_texts[0]:
                            match = re.search(r'^(.+?)\s*\(([^)]+)\)\s*$', cell_texts[0])
                            if match:
                                content = match.group(1).strip()
                                category = match.group(2).strip()
                            else:
                                content = cell_texts[0]
                                if len(cell_texts) > 1:
                                    category = cell_texts[1].strip()
                        
                        for cell_text in cell_texts:
                            matches = re.findall(r'\(([^)]+)\)', cell_text)
                            if matches:
                                cat = matches[-1].strip()
                                if "срок" not in cat.lower() and "выполнен" not in cat.lower():
                                    if not category:
                                        category = cat
                                    clean = re.sub(r'\s*\([^)]+\)\s*$', '', cell_text).strip()
                                    if clean and not content:
                                        content = clean
                        
                        if content:
                            all_contents.append(content)
                            if category:
                                content_to_category[content.lower()] = category.lower()
                
                # Обновляем маппинг
                if content_to_category:
                    # Преобразуем категории в формат для бланка сведений
                    category_map_to_blank = {
                        "канализация": "КАНАЛИЗАЦИЯ",
                        "водопровод": "ВОДОПРОВОД",
                        "водоотведение": "КАНАЛИЗАЦИЯ",
                        "водоснабжение": "ВОДОПРОВОД",
                        "перекладка": "ПЕРЕКЛАДКА",
                        "в/колонки": "В/КОЛОНКИ",
                        "колонки": "В/КОЛОНКИ",
                        "пожарные гидранты": "ПОЖАРНЫЕ ГИДРАНТЫ",
                        "гидранты": "ПОЖАРНЫЕ ГИДРАНТЫ",
                        "частные врезки": "ЧАСТНЫЕ ВРЕЗКИ",
                        "врезки": "ЧАСТНЫЕ ВРЕЗКИ",
                    }
                    
                    problem_to_blank = {}
                    for content, cat in content_to_category.items():
                        # Преобразуем категорию
                        blank_cat = category_map_to_blank.get(cat, "ВОДОПРОВОД")
                        # Пробуем найти по ключевым словам
                        for key, value in category_map_to_blank.items():
                            if key in cat:
                                blank_cat = value
                                break
                        problem_to_blank[content] = blank_cat
                    
                    default_mapping.update(problem_to_blank)
                    
                    # Обновляем problem_categories, группируя по категориям
                    # Создаем временные списки
                    vodootvedenie = []
                    vodsnabshenie = []
                    obshie = []
                    
                    for content in set(all_contents):
                        content_lower = content.lower()
                        cat = content_to_category.get(content_lower, "")
                        
                        # Определяем в какую категорию поместить
                        if "канализац" in cat or "водоотвед" in cat:
                            vodootvedenie.append(content)
                        elif "водопровод" in cat or "водоснабж" in cat or "вод" in cat:
                            vodsnabshenie.append(content)
                        else:
                            obshie.append(content)
                    
                    # Обновляем только если нашли данные
                    if vodootvedenie or vodsnabshenie or obshie:
                        default_categories["водоотведение"] = sorted(set(vodootvedenie + default_categories["водоотведение"]), key=lambda s: s.lower())
                        default_categories["водоснабжение"] = sorted(set(vodsnabshenie + default_categories["водоснабжение"]), key=lambda s: s.lower())
                        default_categories["общие"] = sorted(set(obshie + default_categories["общие"]), key=lambda s: s.lower())
                
            except Exception as e:
                # Если ошибка при чтении, используем значения по умолчанию
                pass
        
        return default_categories, default_mapping
    
    # Загружаем категории из Word документа
    problem_categories, problem_to_blank_category_raw = load_problem_categories_from_docx()
    
    # Преобразуем problem_to_blank_category_raw в правильный формат
    problem_to_blank_category = {}
    category_map_to_blank = {
        "канализация": "КАНАЛИЗАЦИЯ",
        "водопровод": "ВОДОПРОВОД",
        "водоотведение": "КАНАЛИЗАЦИЯ",
        "водоснабжение": "ВОДОПРОВОД",
        "перекладка": "ПЕРЕКЛАДКА",
        "в/колонки": "В/КОЛОНКИ",
        "колонки": "В/КОЛОНКИ",
        "пожарные гидранты": "ПОЖАРНЫЕ ГИДРАНТЫ",
        "гидранты": "ПОЖАРНЫЕ ГИДРАНТЫ",
        "частные врезки": "ЧАСТНЫЕ ВРЕЗКИ",
        "врезки": "ЧАСТНЫЕ ВРЕЗКИ",
    }
    
    for content, cat in problem_to_blank_category_raw.items():
        blank_cat = category_map_to_blank.get(cat, "ВОДОПРОВОД")
        for key, value in category_map_to_blank.items():
            if key in cat.lower():
                blank_cat = value
                break
        problem_to_blank_category[content] = blank_cat
    
    def clean_problem_text(problem_text: str) -> str:
        """
        Удаляет категорию из скобок из текста проблемы.
        Например: "Течь канализации по дороге (канализация)" -> "Течь канализации по дороге"
        Но сохраняет "срок выполнения" в скобках.
        """
        if not problem_text:
            return ""
        import re
        # Сначала сохраняем "срок выполнения" если он есть
        deadline_match = re.search(r'\([^)]*срок[^)]*выполнен[^)]*\)', problem_text, flags=re.IGNORECASE)
        deadline_text = deadline_match.group(0) if deadline_match else ""
        # Удаляем все скобки (включая категории)
        cleaned = re.sub(r'\s*\([^)]*\)', '', problem_text)
        # Восстанавливаем "срок выполнения" если был
        if deadline_text:
            cleaned = cleaned.strip() + " " + deadline_text
        return cleaned.strip()
    
    # Все проблемы для удобства (очищенные от категорий в скобках)
    all_problems = []
    for cat_problems in problem_categories.values():
        # Очищаем каждую проблему от категорий в скобках для отображения
        cleaned_problems = [clean_problem_text(p) for p in cat_problems]
        all_problems.extend(cleaned_problems)

    ttk.Label(frame_add, text="Категория:").grid(row=2, column=0, sticky=tk.W, pady=5)
    category_var = tk.StringVar(value="")
    category_combo = ttk.Combobox(frame_add, textvariable=category_var, values=["", "Водоотведение", "Водоснабжение"], state="readonly", width=37)
    category_combo.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

    ttk.Label(frame_add, text="Содержание заявки:").grid(row=3, column=0, sticky=tk.W, pady=5)
    problem_var = tk.StringVar()
    
    def update_problem_options(*args):
        """Обновляет список проблем в зависимости от выбранной категории."""
        selected_category = category_var.get().lower()
        if selected_category == "водоотведение":
            filtered_options = problem_categories["водоотведение"]
        elif selected_category == "водоснабжение":
            filtered_options = problem_categories["водоснабжение"]
        else:
            # Если категория не выбрана, показываем только общие
            filtered_options = problem_categories["общие"]
        
        problem_entry['values'] = sorted(filtered_options, key=lambda s: s.lower())
        # Сбрасываем выбранное значение, если оно не входит в новый список
        current_value = problem_var.get()
        if current_value and current_value not in filtered_options:
            problem_var.set("")
    
    category_var.trace('w', update_problem_options)
    
    # Инициализация списка проблем (по умолчанию - общие)
    problem_entry = ttk.Combobox(frame_add, textvariable=problem_var, values=sorted(problem_categories["общие"], key=lambda s: s.lower()), state="readonly", width=37)
    problem_entry.grid(row=3, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

    ttk.Label(frame_add, text="Номер бригады:").grid(row=4, column=0, sticky=tk.W, pady=5)
    brigade_var = tk.StringVar()
    brigade_entry = ttk.Entry(frame_add, textvariable=brigade_var, width=37)
    brigade_entry.grid(row=4, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

    ttk.Label(frame_add, text="Срок выполнения (часы):").grid(row=5, column=0, sticky=tk.W, pady=5)
    deadline_var = tk.StringVar(value="")
    deadline_entry = ttk.Entry(frame_add, textvariable=deadline_var, width=37)
    deadline_entry.grid(row=5, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

    ttk.Label(frame_add, text="Телефон звонившего:").grid(row=6, column=0, sticky=tk.W, pady=5)
    phone_entry = ttk.Entry(frame_add, width=40)
    phone_entry.grid(row=6, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

    ttk.Label(frame_add, text="Адрес:").grid(row=7, column=0, sticky=tk.W, pady=5)
    address_entry = ttk.Entry(frame_add, width=40)
    address_entry.grid(row=7, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

    ttk.Label(frame_add, text="Постановка на выполнение:").grid(row=8, column=0, sticky=tk.W, pady=5)
    assignment_entry = DateEntry(frame_add, date_pattern='dd.MM.yyyy', width=15)
    assignment_entry.grid(row=8, column=1, sticky=tk.W, padx=(10, 0), pady=5)

    ttk.Label(frame_add, text="Состояние выполнения:").grid(row=9, column=0, sticky=tk.W, pady=5)
    status_var = tk.StringVar(value="в работе")
    status_combo = ttk.Combobox(frame_add, textvariable=status_var, values=["в работе", "не выполнено", "выполнено"], state="readonly", width=37)
    status_combo.grid(row=9, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

    ttk.Label(frame_add, text="Примечание:").grid(row=10, column=0, sticky=tk.W, pady=5)
    improvement_entry = ttk.Entry(frame_add, width=40)
    improvement_entry.grid(row=10, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

    def add_record():
        name = name_entry.get()
        surname = surname_entry.get()
        problem_text = problem_entry.get()
        deadline_hours = deadline_entry.get().strip()

        # Если срок выполнения указан — провалидируем как положительное число
        if deadline_hours:
            try:
                deadline_int = int(deadline_hours)
                if deadline_int <= 0:
                    messagebox.showwarning("Ошибка", "Срок выполнения должен быть положительным числом")
                    return
            except ValueError:
                messagebox.showwarning("Ошибка", "Срок выполнения должен быть числом")
                return

        # Очищаем текст проблемы от категорий в скобках (если они есть)
        problem_text_cleaned = clean_problem_text(problem_text) if problem_text else ""
        
        # Включаем срок в текст проблемы только если он указан
        if problem_text_cleaned:
            problem = f"{problem_text_cleaned} (срок выполнения {deadline_hours} ч)" if deadline_hours else problem_text_cleaned
        else:
            problem = ""
        phone = phone_entry.get()
        address = address_entry.get()
        date = datetime.now().strftime("%Y-%m-%d")
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            assignment_date = assignment_entry.get_date().strftime("%Y-%m-%d") if assignment_entry.get() else ""
        except Exception:
            assignment_date = ""
        # Статус при создании фиксируем с отметкой времени
        base_status_value = status_var.get()
        status_value = f"{base_status_value} ({datetime.now().strftime('%d.%m.%Y %H:%M')})"
        improvement = improvement_entry.get()
        # Обработка номера бригады: добавляем ".бр" если номер указан
        brigade_number = brigade_var.get().strip()
        if brigade_number:
            if not brigade_number.endswith(".бр"):
                brigade_number = f"{brigade_number}.бр"
        category = category_var.get()
        
        # Проверяем, нужно ли сбросить ID для нового года
        current_year = datetime.now().year
        try:
            # Проверяем максимальный ID за текущий год
            cursor.execute("""
                SELECT MAX(id) FROM records 
                WHERE strftime('%Y', date) = ?
            """, (str(current_year),))
            max_id_result = cursor.fetchone()
            max_id_this_year = max_id_result[0] if max_id_result and max_id_result[0] else 0
            
            # Если это первая запись в году (max_id_this_year == 0),
            # нужно сбросить sequence для таблицы records
            if max_id_this_year == 0:
                # Проверяем, есть ли записи за прошлые годы
                cursor.execute("SELECT MAX(id) FROM records")
                max_id_all = cursor.fetchone()[0] if cursor.fetchone() else 0
                
                # Если есть записи за прошлые годы, сбрасываем sequence
                if max_id_all > 0:
                    try:
                        # Удаляем запись из sqlite_sequence для таблицы records
                        cursor.execute("DELETE FROM sqlite_sequence WHERE name='records'")
                        conn.commit()
                    except Exception:
                        # Если таблица sqlite_sequence не существует или произошла ошибка, игнорируем
                        pass
        except Exception:
            pass
        
        # Получаем или создаём category_id в справочнике категорий
        category_id = None
        if category:
            try:
                cursor.execute("INSERT OR IGNORE INTO common.categories (name) VALUES (?)", (category,))
                cursor.execute("SELECT id FROM common.categories WHERE name=?", (category,))
                row_cat = cursor.fetchone()
                category_id = row_cat[0] if row_cat else None
            except Exception:
                pass

        ts_now = datetime.now().strftime('%d.%m.%Y %H:%M')
        deadline_hours_int = int(deadline_hours) if deadline_hours and str(deadline_hours).isdigit() else None
        cursor.execute(
            "INSERT INTO records "
            "(name, surname, problem, problem_text, deadline_hours, phone, address, date, assignment_date, "
            " status, status_clean, status_changed_at, "
            " user_id, created_at, improvement, brigade_number, category, category_id) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (name, surname, problem, problem_text_cleaned, deadline_hours_int,
             phone, address, date, assignment_date,
             status_value, base_status_value, ts_now,
             user[0], created_at, improvement, brigade_number, category, category_id)
        )
        new_id = cursor.lastrowid
        conn.commit()
        try:
            show_auto_close_info(f"Заявка создана. Порядковый номер: {new_id}\nТелефон: {phone}", duration_ms=8000)
        except Exception:
            pass
        try:
            refresh_records_default()
        except Exception:
            pass
        try:
            refresh_recent()
        except Exception:
            pass
        name_entry.delete(0, tk.END)
        surname_entry.delete(0, tk.END)
        category_var.set("")
        problem_var.set("")
        brigade_var.set("")
        deadline_var.set("")
        phone_entry.delete(0, tk.END)
        address_entry.delete(0, tk.END)
        status_var.set("в работе")
        improvement_entry.delete(0, tk.END)

    ttk.Button(frame_add, text="Добавить", command=add_record).grid(row=11, column=0, columnspan=2, pady=10)

    # ---- Последние заявки  ----
    problem_filter_var = tk.StringVar(value="")
    status_filter_var = tk.StringVar(value="")
    operator_filter_var = tk.StringVar(value="")

    frame_recent = ttk.LabelFrame(main, text="Последние заявки", padding="12")
    frame_recent.grid(row=2, column=0, columnspan=2, padx=16, pady=(0, 0), sticky="nsew")
    frame_recent.rowconfigure(0, weight=1)
    frame_recent.columnconfigure(0, weight=1)

    columns_recent = ("ID", "Наименование", "Фамилия", "Содержание заявки", "Телефон", "Адрес", "Дата", "Постановка", "Состояние", "Пользователь")
    recent_tree = ttk.Treeview(frame_recent, columns=columns_recent, show='headings', height=4)
    for col in columns_recent:
        recent_tree.heading(col, text=col)
        if col == "ID":
            recent_tree.column(col, width=70, minwidth=60, stretch=False)
        elif col in ("Дата", "Постановка", "Состояние"):
            recent_tree.column(col, width=120, minwidth=100, stretch=False)
        elif col in ("Наименование", "Фамилия", "Пользователь"):
            recent_tree.column(col, width=120, minwidth=100, stretch=False)
        elif col == "Телефон":
            recent_tree.column(col, width=120, minwidth=100, stretch=False)
        else:
            recent_tree.column(col, width=220, minwidth=140, stretch=True)
    scrollbar_recent_x = ttk.Scrollbar(frame_recent, orient="horizontal", command=recent_tree.xview)
    scrollbar_recent_y = ttk.Scrollbar(frame_recent, orient="vertical", command=recent_tree.yview)
    recent_tree.configure(xscrollcommand=scrollbar_recent_x.set, yscrollcommand=scrollbar_recent_y.set)
    recent_tree.grid(row=0, column=0, sticky="nsew")
    scrollbar_recent_y.grid(row=0, column=1, sticky="ns")
    scrollbar_recent_x.grid(row=1, column=0, sticky="ew")
    frame_recent.rowconfigure(0, weight=1)

    def refresh_recent():
        # Показываем три последних по id
        for item in recent_tree.get_children():
            recent_tree.delete(item)
        try:
            cursor.execute(
                """
                SELECT * FROM (
                    SELECT r.id, r.name, r.surname, r.problem, r.phone, r.address,
                        COALESCE(r.created_at, r.date) AS created_at,
                        r.assignment_date, r.status, u.username
                    FROM records r
                    LEFT JOIN common.users u ON r.user_id = u.id
                    ORDER BY r.id DESC
                    LIMIT 11
                ) sub
                ORDER BY sub.id ASC;

                """
            )
            rows = cursor.fetchall()
        except Exception:
            rows = []
        for row in rows:
            rec_id, name_v, surname_v, problem_v, phone_v, address_v, dt_str, assign_str, status_text, user_v = row
            # Форматируем дату
            try:
                dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
                date_fmt = dt.strftime("%d.%m.%Y %H:%M")
            except Exception:
                try:
                    d = datetime.strptime(dt_str, "%Y-%m-%d")
                    date_fmt = d.strftime("%d.%m.%Y")
                except Exception:
                    date_fmt = dt_str or ""
            # Форматируем дату постановки
            assignment_formatted = ""
            if assign_str:
                try:
                    assignment_formatted = datetime.strptime(assign_str, "%Y-%m-%d").strftime("%d.%m.%Y")
                except Exception:
                    assignment_formatted = assign_str
            recent_tree.insert('', 'end', values=(
                rec_id,
                name_v or "",
                surname_v or "",
                problem_v or "",
                phone_v or "",
                address_v or "",
                date_fmt,
                assignment_formatted,
                status_text or "",
                user_v or "",
            ))
    refresh_recent()

    # Подпись состояния базы для автообновления (max(id), count(*))
    last_seen_recent_sig = None
    def _get_recent_signature():
        try:
            cursor.execute("SELECT COALESCE(MAX(id),0), COUNT(*) FROM records")
            row = cursor.fetchone()
            return (row[0] or 0, row[1] or 0)
        except Exception:
            return None
    try:
        last_seen_recent_sig = _get_recent_signature()
    except Exception:
        last_seen_recent_sig = None

    # Состояние отдельного окна со списком
    records_window = None
    records_tree = None

    def _execute_query_multiple_dbs(query_func, db_files=None):
        """
        Выполняет запрос к нескольким базам данных и объединяет результаты.
        query_func - функция, которая принимает cursor и возвращает результаты запроса.
        db_files - список путей к базам. Если None, используется текущая база.
        """
        if db_files is None:
            # Используем только текущую базу
            return query_func(cursor)
        
        all_results = []
        original_db = current_db_file
        original_conn = conn
        original_cursor = cursor
        
        try:
            for db_file in db_files:
                db_path = str(db_file)
                if not os.path.exists(db_path):
                    continue
                
                # Подключаемся к базе
                temp_conn = sqlite3.connect(db_path)
                temp_cursor = temp_conn.cursor()
                
                # Обновляем схему базы данных (добавляем отсутствующие колонки)
                try:
                    ensure_schema_for_connection(temp_conn, temp_cursor)
                except Exception as e:
                    print(f"Предупреждение: не удалось обновить схему для {db_path}: {e}")
                
                # Присоединяем общую базу пользователей
                try:
                    _common_path = Path(_get_common_db_path())
                    _common_escaped = str(_common_path).replace("'", "''")
                    temp_conn.execute(f"ATTACH DATABASE '{_common_escaped}' AS common")
                except Exception:
                    pass
                
                try:
                    # Выполняем запрос
                    results = query_func(temp_cursor)
                    all_results.extend(results)
                except Exception as e:
                    print(f"Ошибка при запросе к {db_path}: {e}")
                finally:
                    temp_conn.close()
            
            return all_results
        finally:
            # Восстанавливаем оригинальное подключение
            pass  # Не переключаем обратно, так как это может нарушить работу интерфейса

    def _get_all_databases():
        """Возвращает список всех баз данных (кроме app.db)."""
        base_dir = _load_db_base_dir()
        db_files = []
        for db_file in Path(base_dir).glob("app_*.db"):
            if db_file.name != "app.db":  # Исключаем общую базу
                db_files.append(db_file)
        # Сортируем по дате (из имени файла)
        db_files.sort(key=lambda x: x.name, reverse=True)
        return db_files

    def get_filtered_rows(keyword=None, start_date: str | None = None, end_date: str | None = None, db_files=None):
        """Возвращает строки по текущим фильтрам и опциональному диапазону дат (из всех баз данных)."""
        # Используем selected_db_files если db_files не указан
        if db_files is None:
            try:
                db_files = selected_db_files
            except:
                pass
        
        def execute_query(cursor_to_use):
            query = """SELECT r.id, r.name, r.surname, r.category, r.problem, r.brigade_number, r.phone, r.address,
                               COALESCE(r.created_at, r.date) AS created_at,
                               r.assignment_date, r.status, u.username 
                       FROM records r 
                       LEFT JOIN common.users u ON r.user_id = u.id"""
            where_clauses = []
            params = []
            if keyword:
                _kw = (keyword or "").lower()
                where_clauses.append("(LOWER(r.name) LIKE ? OR LOWER(r.surname) LIKE ? OR LOWER(r.problem) LIKE ? OR LOWER(r.phone) LIKE ? OR LOWER(r.address) LIKE ? OR LOWER(u.username) LIKE ?)")
                params.extend([f"%{_kw}%", f"%{_kw}%", f"%{_kw}%", f"%{_kw}%", f"%{_kw}%", f"%{_kw}%"])
                # Поиск по ID (точное совпадение, если введены только цифры)
                try:
                    if str(keyword).isdigit():
                        where_clauses.append("r.id = ?")
                        params.append(int(keyword))
                except Exception:
                    pass
            if problem_filter_var.get():
                _pv = problem_filter_var.get().lower()
                where_clauses.append("LOWER(r.problem) LIKE ?")
                params.append(f"%{_pv}%")
            if status_filter_var.get():
                # Статус хранится с отметкой времени, фильтруем по префиксу
                where_clauses.append("r.status LIKE ?")
                params.append(status_filter_var.get() + "%")
            if operator_filter_var.get():
                where_clauses.append("u.username = ?")
                params.append(operator_filter_var.get())
            if start_date and end_date:
                where_clauses.append("date(r.date) >= date(?) AND date(r.date) <= date(?)")
                params.extend([start_date, end_date])
            if where_clauses:
                query += " WHERE " + " AND ".join(where_clauses)
            cursor_to_use.execute(query, params)
            return cursor_to_use.fetchall()
        
        # Используем selected_db_files если указан, иначе все базы
        if db_files is None:
            try:
                db_files = selected_db_files
            except:
                pass
            if db_files is None:
                # Если selected_db_files не установлен, используем только текущую базу
                return execute_query(cursor)
        
        # Используем функцию для запроса к нескольким базам
        return _execute_query_multiple_dbs(execute_query, db_files)

    # Локальный ненавязчивый тост без кнопок, закрывается сам
    def show_auto_close_info(message_text: str, duration_ms: int = 8000):
        try:
            toast = tk.Toplevel(main)
            setup_scaling()
            toast.overrideredirect(True)
            try:
                toast.attributes("-topmost", True)
            except Exception:
                pass
            # Полностью белое окно и центрированный текст
            toast.configure(bg="#ffffff")
            inner = tk.Frame(toast, bg="#ffffff")
            inner.pack(fill="both", expand=True)
            lbl = tk.Label(inner, text=message_text, font=("Segoe UI", 16), bg="#ffffff", fg="#111827")
            lbl.pack(expand=True, padx=28, pady=28)
            try:
                lbl.configure(wraplength=560, justify="center")
            except Exception:
                pass

            # Автоподгон и позиционирование по центру главного окна
            try:
                toast.update_idletasks()
                w = max(480, toast.winfo_reqwidth())
                h = max(140, toast.winfo_reqheight())
                # Позиционируем в правом верхнем углу, чтобы не перекрывать центральные элементы
                x0 = main.winfo_rootx()
                y0 = main.winfo_rooty()
                mw = main.winfo_width() or 800
                x = x0 + mw - w - 24
                y = y0 + 24
                toast.geometry(f"{w}x{h}+{x}+{y}")
            except Exception:
                pass

            # Лёгкая прозрачность и авто-закрытие, без перехвата кликов
            try:
                toast.attributes('-alpha', 0.98)
            except Exception:
                pass
            toast.after(duration_ms, toast.destroy)
        except Exception:
            # Фолбэк: обычное сообщение, если что-то пошло не так
            try:
                messagebox.showinfo("", message_text)
            except Exception:
                pass

    def get_filtered_rows_for(keyword: str | None, problem_value: str | None, status_value: str | None,
                               operator_value: str | None, start_date: str | None = None, end_date: str | None = None):
        """Возвращает строки по параметрам (для фильтров в окне списка) из всех баз данных."""
        def execute_query(cursor_to_use):
            query = """SELECT r.id, r.name, r.surname, r.category, r.problem, r.brigade_number, r.phone, r.address,
                               COALESCE(r.created_at, r.date) AS created_at,
                               r.assignment_date, r.status, u.username 
                       FROM records r 
                       LEFT JOIN common.users u ON r.user_id = u.id"""
            where_clauses = []
            params = []
            # Поиск по ключевому слову
            if keyword:
                # Если поиск по ID - используем точный поиск
                if str(keyword).isdigit():
                    where_clauses.append("r.id = ?")
                    params.append(int(keyword))
                else:
                    # Для текстового поиска - получаем все записи и фильтруем в Python
                    # Это решает проблему с кириллицей в SQLite
                    where_clauses.append("1=1")
            if problem_value:
                where_clauses.append("r.problem LIKE ?")
                params.append(f"%{problem_value}%")
            if status_value:
                # Статус хранится с отметкой времени, фильтруем по префиксу
                where_clauses.append("r.status LIKE ?")
                params.append(status_value + "%")
            if operator_value:
                where_clauses.append("u.username = ?")
                params.append(operator_value)
            if start_date and end_date:
                where_clauses.append("date(r.date) >= date(?) AND date(r.date) <= date(?)")
                params.extend([start_date, end_date])
            if where_clauses:
                query += " WHERE " + " AND ".join(where_clauses)
            cursor_to_use.execute(query, params)
            return cursor_to_use.fetchall()
        
        # Получаем все базы данных
        db_files = _get_all_databases()
        if not db_files:
            # Если баз нет, используем текущую
            results = execute_query(cursor)
        else:
            # Используем функцию для запроса к нескольким базам
            results = _execute_query_multiple_dbs(execute_query, db_files)
        
        # Если есть текстовый поиск (не по ID), фильтруем результаты в Python
        if keyword and not str(keyword).isdigit():
            keyword_lower = keyword.lower()
            filtered_results = []
            for row in results:
                # Проверяем все текстовые поля на совпадение
                # row[0]=id, row[1]=name, row[2]=surname, row[3]=category, row[4]=problem, row[5]=brigade_number, row[6]=phone, row[7]=address, row[8]=created_at, row[9]=assignment_date, row[10]=status, row[11]=username
                name = (row[1] or "").lower()
                surname = (row[2] or "").lower()
                category = (row[3] or "").lower()
                problem = (row[4] or "").lower()
                phone = (row[6] or "").lower()
                address = (row[7] or "").lower()
                username = (row[11] or "").lower()
                
                if (keyword_lower in name or keyword_lower in surname or 
                    keyword_lower in category or keyword_lower in problem or keyword_lower in phone or 
                    keyword_lower in address or keyword_lower in username):
                    filtered_results.append(row)
            return filtered_results
        
        return results

    def populate_tree_from_rows(rows):
        nonlocal records_tree
        if not records_tree or not records_tree.winfo_exists():
            return
        for item in records_tree.get_children():
            records_tree.delete(item)
        
        # Группируем записи по месяцам для разграничения (только если выбрано несколько баз)
        current_month_label = None
        months_ru = {
            "January": "Январь", "February": "Февраль", "March": "Март",
            "April": "Апрель", "May": "Май", "June": "Июнь",
            "July": "Июль", "August": "Август", "September": "Сентябрь",
            "October": "Октябрь", "November": "Ноябрь", "December": "Декабрь"
        }
        
        for i, row in enumerate(rows):
            # Определяем месяц записи (только если выбрано несколько баз)
            month_label = None
            if selected_db_files and len(selected_db_files) > 1:
                try:
                    dt = datetime.strptime(row[8], "%Y-%m-%d %H:%M:%S")
                    month_label = dt.strftime("%B %Y")
                except Exception:
                    try:
                        d = datetime.strptime(row[8], "%Y-%m-%d")
                        month_label = d.strftime("%B %Y")
                    except Exception:
                        pass
                
                # Преобразуем месяц на русский
                if month_label:
                    for en, ru in months_ru.items():
                        month_label = month_label.replace(en, ru)
                
                # Добавляем заголовок месяца если изменился
                if month_label and month_label != current_month_label:
                    current_month_label = month_label
                    # Вставляем строку-разделитель с названием месяца
                    records_tree.insert('', 'end', values=(
                        "", "", f"═══════ {month_label.upper()} ═══════", "", "", "", "", "", "", "", ""
                    ), tags=('header',))
            
            # row[8] может содержать дату или дату-время (created_at)
            date_formatted = ""; time_formatted = ""
            try:
                dt = datetime.strptime(row[8], "%Y-%m-%d %H:%M:%S")
                date_formatted = dt.strftime("%d.%m.%Y")
                time_formatted = dt.strftime("%H:%M")
            except Exception:
                try:
                    d = datetime.strptime(row[8], "%Y-%m-%d")
                    date_formatted = d.strftime("%d.%m.%Y")
                except Exception:
                    date_formatted = str(row[8] or "")
            assignment_formatted = ""
            if row[9]:
                try:
                    assignment_formatted = datetime.strptime(row[9], "%Y-%m-%d").strftime("%d.%m.%Y")
                except:
                    assignment_formatted = row[9]
            tag = 'even' if i % 2 == 0 else 'odd'
            records_tree.insert('', 'end', values=(
                row[0], row[1], row[2], row[3] or "", row[4] or "", row[5] or "", row[6] or "",
                row[7] or "", date_formatted, time_formatted, assignment_formatted, row[10] or "", row[11] or ""
            ), tags=(tag,))

    def refresh_records_default():
        """Обновляет окно списка по текущим фильтрам (без строки поиска)."""
        rows = get_filtered_rows(None)
        populate_tree_from_rows(rows)

    def apply_filters():
        nonlocal records_window, records_tree
        if not records_window or not records_window.winfo_exists():
            open_records_window()
        # Без ограничения по датам; для дат используйте кнопку "Поиск по датам"
        rows = get_filtered_rows(None, None, None)
        populate_tree_from_rows(rows)



    def reset_filters():
        nonlocal records_window, records_tree
        problem_filter_var.set("")
        status_filter_var.set("")
        operator_filter_var.set("")
        if not records_window or not records_window.winfo_exists():
            open_records_window()
        # Показать все записи без ограничений по датам
        rows = get_filtered_rows(None, None, None)
        populate_tree_from_rows(rows)

    # Блок фильтров и поиск по датам на главном экране удалены по требованию


    # ---- Список записей отдельным окном ----
    # Переменная для хранения выбранных баз (None = текущая база, список = несколько баз)
    selected_db_files = None
    
    def open_records_window():
        nonlocal records_window, records_tree, selected_db_files
        
        def choose_db_dialog(btn_choose_db):
            nonlocal selected_db_files
            win = tk.Toplevel(records_window)
            win.title("Выбор базы данных")
            setup_scaling()
            fit_and_center_window(win, min_width=400, min_height=350)
            
            # Радиокнопки для выбора режима
            mode_var = tk.StringVar(value="current")
            ttk.Label(win, text="Режим отображения:", font=("", 10, "bold")).pack(pady=(10, 5))
            
            ttk.Radiobutton(win, text="Текущий месяц (по умолчанию)", variable=mode_var, value="current").pack(anchor=tk.W, padx=20)
            ttk.Radiobutton(win, text="Вся база (все месяцы)", variable=mode_var, value="all").pack(anchor=tk.W, padx=20)
            ttk.Radiobutton(win, text="Выбрать несколько месяцев", variable=mode_var, value="multiple").pack(anchor=tk.W, padx=20)
            
            # Фрейм для выбора одного месяца
            single_month_frame = ttk.Frame(win)
            single_month_frame.pack(pady=10, fill=tk.X, padx=20)
            ttk.Label(single_month_frame, text="Выберите год:").grid(row=0, column=0, sticky=tk.W, pady=2)
            year_var = tk.IntVar(value=datetime.now().year)
            year_spin = ttk.Spinbox(single_month_frame, from_=2020, to=2100, textvariable=year_var, width=8)
            year_spin.grid(row=0, column=1, padx=5, pady=2)
            
            ttk.Label(single_month_frame, text="Выберите месяц:").grid(row=1, column=0, sticky=tk.W, pady=2)
            month_var = tk.IntVar(value=datetime.now().month)
            month_spin = ttk.Spinbox(single_month_frame, from_=1, to=12, textvariable=month_var, width=8)
            month_spin.grid(row=1, column=1, padx=5, pady=2)
            
            def update_ui():
                if mode_var.get() == "current":
                    single_month_frame.pack_forget()
                else:
                    single_month_frame.pack(pady=10, fill=tk.X, padx=20)
            
            mode_var.trace('w', lambda *args: update_ui())
            update_ui()
            
            def do_select():
                nonlocal selected_db_files
                mode = mode_var.get()
                base_dir = _load_db_base_dir()
                
                if mode == "current":
                    # Текущий месяц - возвращаемся к одной базе
                    year = datetime.now().year
                    month = datetime.now().month
                    db_name = f"app_{year}_{month:02d}.db"
                    db_path = str(Path(base_dir) / db_name)
                    if not os.path.exists(db_path):
                        messagebox.showwarning("База не найдена", f"База за {month:02d}.{year} не существует!")
                        return
                    selected_db_files = None
                    switch_db(db_path)
                    btn_choose_db.config(text=f"Сменить базу ({get_month_year_label(os.path.basename(current_db_file))})")
                    
                elif mode == "all":
                    # Вся база - выбираем все базы
                    all_dbs = []
                    for db_file in Path(base_dir).glob("app_*.db"):
                        if db_file.name != "app.db":
                            all_dbs.append(db_file)
                    all_dbs.sort(key=lambda x: x.name)
                    if not all_dbs:
                        messagebox.showwarning("База не найдена", "Базы данных не найдены!")
                        return
                    selected_db_files = all_dbs
                    btn_choose_db.config(text=f"Сменить базу (ВСЯ БАЗА - {len(all_dbs)} месяцев)")
                    
                elif mode == "multiple":
                    # Несколько месяцев - используем существующий диалог
                    db_files = []
                    for db_file in Path(base_dir).glob("app_*.db"):
                        if db_file.name != "app.db":
                            db_files.append(db_file)
                    db_files.sort(key=lambda x: x.name, reverse=True)
                    
                    if not db_files:
                        messagebox.showinfo("Информация", "Базы данных не найдены")
                        return
                    
                    # Создаем диалог выбора
                    select_win = tk.Toplevel(win)
                    select_win.title("Выбор месяцев")
                    fit_and_center_window(select_win, min_width=350, min_height=400)
                    
                    frm = ttk.Frame(select_win, padding="20")
                    frm.pack(fill=tk.BOTH, expand=True)
                    
                    ttk.Label(frm, text="Выберите месяцы:", font=("", 10, "bold")).pack(pady=(0, 10))
                    
                    list_frame = ttk.Frame(frm)
                    list_frame.pack(fill=tk.BOTH, expand=True, pady=5)
                    
                    scrollbar = ttk.Scrollbar(list_frame)
                    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                    
                    listbox = tk.Listbox(list_frame, selectmode=tk.MULTIPLE, yscrollcommand=scrollbar.set, height=12)
                    listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                    scrollbar.config(command=listbox.yview)
                    
                    selected_vars = {}
                    for db_file in db_files:
                        label = get_month_year_label(db_file.name)
                        idx = listbox.size()
                        listbox.insert(tk.END, label)
                        selected_vars[idx] = db_file
                    
                    # Выбираем текущий месяц по умолчанию
                    try:
                        current_db_name = os.path.basename(current_db_file)
                        for idx, db_file in enumerate(db_files):
                            if db_file.name == current_db_name:
                                listbox.selection_set(idx)
                                listbox.see(idx)
                                break
                    except:
                        pass
                    
                    selected_dbs_list = []
                    
                    def confirm_multiple():
                        nonlocal selected_db_files, selected_dbs_list
                        selected_indices = listbox.curselection()
                        if not selected_indices:
                            messagebox.showwarning("Предупреждение", "Выберите хотя бы один месяц")
                            return
                        selected_dbs_list = [selected_vars[idx] for idx in selected_indices]
                        selected_db_files = selected_dbs_list
                        select_win.destroy()
                        win.destroy()
                        btn_choose_db.config(text=f"Сменить базу ({len(selected_dbs_list)} месяцев)")
                        refresh_records_default()
                        try:
                            refresh_recent()
                        except Exception:
                            pass
                    
                    def select_all_multiple():
                        listbox.selection_set(0, tk.END)
                    
                    btn_frame = ttk.Frame(frm)
                    btn_frame.pack(pady=5)
                    ttk.Button(btn_frame, text="Выбрать все", command=select_all_multiple).pack(side=tk.LEFT, padx=5)
                    ttk.Button(btn_frame, text="ОК", command=confirm_multiple).pack(side=tk.LEFT, padx=5)
                    ttk.Button(btn_frame, text="Отмена", command=select_win.destroy).pack(side=tk.LEFT, padx=5)
                    
                    select_win.wait_window()
                    return
                
                refresh_records_default()
                try:
                    refresh_recent()
                except Exception:
                    pass
                win.destroy()

            btn_frame = ttk.Frame(win)
            btn_frame.pack(pady=10)
            ttk.Button(btn_frame, text="Выбрать", command=do_select).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="Отмена", command=win.destroy).pack(side=tk.LEFT, padx=5)


    # ...дальше ваш код (filters_panel, таблица и т.д.)...
        if records_window and records_window.winfo_exists():
            try:
                records_window.lift()
                records_window.focus_force()
            except Exception:
                pass
            return

        # Определяем список проблем для фильтров
        problem_options = [
            "забой канализационного колодца",
            "течь воды",
            "течь в / колонки",
            "течь канализации по дороге",
            "течь из-под земли",
            "течь пожарного гидранта",
            "течь из -под асфальта",
            "течь трассы холодной воды",
            "перекладка",
            "откачка воды из колодца",
            "водопровод",
            "частные врезки",
            "открыт колодец (отсутствие/несоответствие крышки)",
            "дефект водоразборной колонки",
            "ржавая холодная вода в жилом фонде",
            "восстановление асфальтобетонного покрытия",
            "Не работает в / колонка",
            "слабое давление х/в",
            "разрушена плита колодца",
            "восстановить в/колонку",
            "привести к/к в нормативное состояние",
            "привести в / к в нормативное состояние",
            "обвал в/к",
            "обвал к/к",
            "течь в/к",
            "нет х/в"
        ]

        records_window = tk.Toplevel(main)
        setup_scaling()
        records_window.title("Список данных")
        records_window.resizable(True, True)
        # Открываем окно во всю ширину экрана (и высоту, если доступно)
        try:
            records_window.state('zoomed')
        except Exception:
            try:
                # Фолбэк: максимально широкое окно с ограничением по высоте
                fit_and_center_window(records_window, min_width=1100, min_height=600, max_ratio=0.98)
                sw = records_window.winfo_screenwidth()
                sh = records_window.winfo_screenheight()
                records_window.geometry(f"{sw}x{int(sh*0.9)}+0+0")
            except Exception:
                pass
        
        # --- Группа действий сверху слева (смена базы + папка экспорта) ---
        export_base_dir_var = tk.StringVar(value="")

        def _choose_export_dir():
            try:
                d = filedialog.askdirectory(title="Выберите папку для экспорта")
            except Exception:
                d = ""
            if d:
                try:
                    export_base_dir_var.set(d)
                    try:
                        show_auto_close_info(f"Папка экспорта: {d}", duration_ms=3500)
                    except Exception:
                        pass
                except Exception:
                    pass

        header_actions = ttk.Frame(records_window)
        header_actions.grid(row=0, column=0, padx=12, pady=(8, 0), sticky="w")

        btn_choose_db = ttk.Button(header_actions,
            text=f"Сменить базу ({get_month_year_label(os.path.basename(current_db_file))})",
            command=lambda: choose_db_dialog(btn_choose_db))
        btn_choose_db.pack(side="left")

        btn_export_dir_top = ttk.Button(header_actions, text="📁", width=3, padding=6, command=_choose_export_dir)
        btn_export_dir_top.pack(side="left", padx=(6, 0))
        # --- Панель фильтров ---
        filters_panel = ttk.LabelFrame(records_window, text="Поиск и фильтры", padding="10")
        filters_panel.grid(row=1, column=0, padx=12, pady=(6, 4), sticky="ew")
        filters_panel.columnconfigure(1, weight=1)
        filters_panel.columnconfigure(3, weight=1)
        filters_panel.columnconfigure(5, weight=1)

        ttk.Label(filters_panel, text="Поиск:").grid(row=0, column=0, sticky="w", padx=(0, 6))
        local_keyword = ttk.Entry(filters_panel, width=30)
        local_keyword.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 12))

        ttk.Label(filters_panel, text="Проблема:").grid(row=0, column=2, sticky="w", padx=(0, 6))
        local_problem_var = tk.StringVar(value="")
        local_problem = ttk.Combobox(filters_panel, textvariable=local_problem_var,
                                     values=[""] + sorted([v for v in problem_options], key=lambda s: s.lower()), state="readonly", width=60)
        local_problem.grid(row=0, column=3, sticky=(tk.W, tk.E), padx=(0, 12))
        try:
            # Показать больше элементов в выпадающем списке и не обрезать длинные строки
            local_problem.configure(height=min(len([""] + [v for v in problem_options]), 20))
        except Exception:
            pass

        # Поднять окно перед показом выпадающего списка, чтобы список был поверх
        def _raise_for_dropdown(event=None):
            try:
                records_window.lift()
                try:
                    records_window.attributes("-topmost", True)
                    records_window.after(400, lambda: records_window.attributes("-topmost", False))
                except Exception:
                    pass
            except Exception:
                pass
        try:
            local_problem.bind("<Button-1>", _raise_for_dropdown)
            local_problem.bind("<FocusIn>", _raise_for_dropdown)
        except Exception:
            pass

        ttk.Label(filters_panel, text="Оператор:").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=(8, 0))
        local_operator_var = tk.StringVar(value="")
        try:
            cursor.execute("SELECT username FROM users ORDER BY username")
            _operator_opts = [""] + [row[0] for row in cursor.fetchall()]
        except Exception:
            _operator_opts = [""]
        local_operator = ttk.Combobox(filters_panel, textvariable=local_operator_var,
                                      values=_operator_opts, state="readonly", width=24)
        local_operator.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=(8, 0))

        ttk.Label(filters_panel, text="Дата начала:").grid(row=0, column=4, sticky="w", padx=(0, 6))
        local_start = DateEntry(filters_panel, date_pattern='dd.MM.yyyy', width=22)
        local_start.grid(row=0, column=5, sticky="w")
        try:
            local_start.delete(0, tk.END)
        except Exception:
            pass

        ttk.Label(filters_panel, text="Дата конца:").grid(row=1, column=4, sticky="w", padx=(12, 6), pady=(8, 0))
        local_end = DateEntry(filters_panel, date_pattern='dd.MM.yyyy', width=22)
        local_end.grid(row=1, column=5, sticky="w", pady=(8, 0))
        try:
            local_end.delete(0, tk.END)
        except Exception:
            pass
        
                # Состояние
        ttk.Label(filters_panel, text="Состояние:").grid(row=1, column=2, sticky="w", padx=(0, 6), pady=(8, 0))
        local_status_var = tk.StringVar(value="")
        local_status = ttk.Combobox(filters_panel, textvariable=local_status_var,
                                    values=["", "в работе", "не выполнено", "выполнено"], state="readonly", width=18)
        local_status.grid(row=1, column=3, sticky=(tk.W, tk.E), pady=(8, 0))

        def apply_local_filters():
            kw = local_keyword.get().strip()
            pv = local_problem_var.get().strip()
            sv = local_status_var.get().strip()
            ov = local_operator_var.get().strip()
            # Получаем даты
            try:
                start_date = local_start.get_date().strftime("%Y-%m-%d") if local_start.get() else None
                end_date = local_end.get_date().strftime("%Y-%m-%d") if local_end.get() else None
            except Exception:
                start_date = None
                end_date = None

            # Если введена только одна дата — используем её и как начало, и как конец
            if start_date and not end_date:
                end_date = start_date
            if end_date and not start_date:
                start_date = end_date

            # Передаём keyword как есть - функция get_filtered_rows_for сама обработает регистр
            rows = get_filtered_rows_for(kw or None, pv or None, sv or None, ov or None, start_date, end_date)
            populate_tree_from_rows(rows)

        def reset_local_filters():
            local_keyword.delete(0, tk.END)
            local_problem_var.set("")
            local_status_var.set("")
            local_operator_var.set("")
            try:
                local_start.delete(0, tk.END)
                local_end.delete(0, tk.END)
            except Exception:
                pass
            rows = get_filtered_rows_for(None, None, None, None, None, None)
            populate_tree_from_rows(rows)

    

        btns_local = ttk.Frame(filters_panel)
        btns_local.grid(row=0, column=6, rowspan=2, sticky="e", padx=(12, 0))
        ttk.Button(btns_local, text="Найти", command=apply_local_filters).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(btns_local, text="Очистить фильтр", command=reset_local_filters).grid(row=0, column=1)
 

        # --- Экспорт из окна списка (всплывающий список) ---
        def _select_multiple_databases():
            """Диалог выбора нескольких баз данных (месяцев). Возвращает список путей к базам или None."""
            base_dir = _load_db_base_dir()
            db_files = []
            
            # Находим все базы данных в директории
            for db_file in Path(base_dir).glob("app_*.db"):
                if db_file.name != "app.db":  # Исключаем общую базу
                    db_files.append(db_file)
            
            if not db_files:
                messagebox.showinfo("Информация", "Базы данных не найдены")
                return None
            
            # Сортируем по дате (из имени файла)
            db_files.sort(key=lambda x: x.name, reverse=True)
            
            # Создаем окно выбора
            select_win = tk.Toplevel(main)
            setup_scaling()
            select_win.title("Выбор баз данных для экспорта")
            select_win.resizable(True, True)
            try:
                fit_and_center_window(select_win, min_width=400, min_height=500)
            except:
                pass
            
            frm = ttk.Frame(select_win, padding="20")
            frm.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            select_win.columnconfigure(0, weight=1)
            select_win.rowconfigure(0, weight=1)
            frm.columnconfigure(0, weight=1)
            frm.rowconfigure(1, weight=1)
            
            ttk.Label(frm, text="Выберите базы данных (месяцы) для экспорта:", font=("", 10, "bold")).grid(row=0, column=0, sticky=tk.W, pady=(0, 10))
            
            # Список с чекбоксами
            list_frame = ttk.Frame(frm)
            list_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
            list_frame.columnconfigure(0, weight=1)
            list_frame.rowconfigure(0, weight=1)
            
            scrollbar = ttk.Scrollbar(list_frame)
            scrollbar.grid(row=0, column=1, sticky="ns")
            
            listbox = tk.Listbox(list_frame, selectmode=tk.MULTIPLE, yscrollcommand=scrollbar.set, height=15)
            listbox.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            scrollbar.config(command=listbox.yview)
            
            # Заполняем список
            selected_vars = {}
            for db_file in db_files:
                label = get_month_year_label(db_file.name)
                listbox.insert(tk.END, label)
                selected_vars[listbox.size() - 1] = db_file
            
            # Выбираем текущую базу по умолчанию
            try:
                current_db_name = os.path.basename(current_db_file)
                for idx, db_file in enumerate(db_files):
                    if db_file.name == current_db_name:
                        listbox.selection_set(idx)
                        listbox.see(idx)
                        break
            except:
                pass
            
            selected_dbs = []
            
            def confirm_selection():
                nonlocal selected_dbs
                selected_indices = listbox.curselection()
                if not selected_indices:
                    messagebox.showwarning("Предупреждение", "Выберите хотя бы одну базу данных")
                    return
                
                selected_dbs = [selected_vars[idx] for idx in selected_indices]
                select_win.destroy()
            
            def select_all():
                listbox.selection_set(0, tk.END)
            
            def deselect_all():
                listbox.selection_clear(0, tk.END)
            
            btn_frame = ttk.Frame(frm)
            btn_frame.grid(row=2, column=0, pady=10)
            ttk.Button(btn_frame, text="Выбрать все", command=select_all).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="Снять выбор", command=deselect_all).pack(side=tk.LEFT, padx=5)
            
            btn_frame2 = ttk.Frame(frm)
            btn_frame2.grid(row=3, column=0, pady=5)
            ttk.Button(btn_frame2, text="Экспортировать", command=confirm_selection).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame2, text="Отмена", command=select_win.destroy).pack(side=tk.LEFT, padx=5)
            
            select_win.wait_window()
            return selected_dbs if selected_dbs else None

        def _db_files_for_date_range(start_date, end_date):
            """Автоматически определяет помесячные базы данных, покрывающие период [start_date, end_date].

            Базы хранятся по месяцам (app_YYYY_MM.db), а отчёты за сутки/неделю оперируют
            локальной датой окна, поэтому нужную базу (или базы, если период затрагивает
            границу месяца) можно вычислить напрямую, не спрашивая пользователя.
            """
            files = []
            seen = set()
            cur = start_date.replace(day=1)
            end_month = end_date.replace(day=1)
            while cur <= end_month:
                path = get_db_full_path_by_date(cur.strftime("%Y-%m-%d"))
                if path not in seen and os.path.exists(path):
                    seen.add(path)
                    files.append(Path(path))
                cur = (cur.replace(year=cur.year + 1, month=1) if cur.month == 12
                       else cur.replace(month=cur.month + 1))
            return files

        def _fetch_rows_for_local(keyword: str | None, problem: str | None, status: str | None, operator: str | None, start_date: str | None, end_date: str | None, db_files=None):
            """Возвращает строки для экспорта по текущим локальным фильтрам окна списка."""
            def execute_query(cursor_to_use):
                query = (
                    """SELECT r.id, r.problem, r.phone, r.address, r.improvement, r.brigade_number, r.category, COALESCE(r.created_at, r.date), r.name, r.surname, u.username, r.status
                        FROM records r
                        LEFT JOIN common.users u ON r.user_id = u.id"""
                )
                where_clauses = []
                params: list[str] = []
                if keyword:
                    _kw = (keyword or "").lower()
                    where_clauses.append("(LOWER(r.name) LIKE ? OR LOWER(r.surname) LIKE ? OR LOWER(r.problem) LIKE ? OR LOWER(r.phone) LIKE ? OR LOWER(r.address) LIKE ? OR LOWER(u.username) LIKE ?)")
                    params.extend([f"%{_kw}%", f"%{_kw}%", f"%{_kw}%", f"%{_kw}%", f"%{_kw}%", f"%{_kw}%"])
                    try:
                        if str(keyword).isdigit():
                            where_clauses.append("r.id = ?")
                            params.append(int(keyword))
                    except Exception:
                        pass
                if problem:
                    _pv = (problem or "").lower()
                    where_clauses.append("LOWER(r.problem) LIKE ?")
                    params.append(f"%{_pv}%")
                if status:
                    # Статус хранится с отметкой времени, фильтруем по префиксу
                    where_clauses.append("r.status LIKE ?")
                    params.append(status + "%")
                if operator:
                    where_clauses.append("u.username = ?")
                    params.append(operator)
                if start_date and end_date:
                    where_clauses.append("date(r.date) >= date(?) AND date(r.date) <= date(?)")
                    params.extend([start_date, end_date])
                if where_clauses:
                    query += " WHERE " + " AND ".join(where_clauses)
                query += " ORDER BY r.id ASC"
                cursor_to_use.execute(query, params)
                return cursor_to_use.fetchall()
            
            return _execute_query_multiple_dbs(execute_query, db_files)

        def _export_current_list():
            """Экспортирует текущий список по локальным фильтрам в отдельный Excel."""
            if not PANDAS_AVAILABLE:
                messagebox.showerror("Ошибка", "Модуль pandas не установлен.\n\n"
                                          "Установите зависимости:\n"
                                          "pip install pandas openpyxl\n\n"
                                          "Или запустите install_dependencies.bat")
                return
            try:
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Alignment
            except ImportError as e:
                messagebox.showerror("Ошибка", f"Не установлен модуль openpyxl: {e}\n\npip install openpyxl")
                return

            # Выбор баз данных
            db_files = _select_multiple_databases()
            if db_files is None:
                return  # Пользователь отменил выбор

            # Используем локальные даты, если заполнены обе
            try:
                start_date = local_start.get_date().strftime("%Y-%m-%d") if local_start.get() else None
                end_date = local_end.get_date().strftime("%Y-%m-%d") if local_end.get() else None
            except Exception:
                start_date = None
                end_date = None

            try:
                rows = _fetch_rows_for_local(
                    local_keyword.get().strip() or None,
                    local_problem_var.get().strip() or None,
                    local_status_var.get().strip() or None,
                    local_operator_var.get().strip() or None,
                    start_date,
                    end_date,
                    db_files,
                )
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при получении данных: {e}")
                return

            if not rows:
                messagebox.showinfo("Результат", "Записей по текущим фильтрам нет")
                return

            formatted_rows = []
            for idx, row in enumerate(rows, start=1):
                rec_id, problem, phone, address, improvement, brigade_number, category, dt_str, applicant_name, applicant_surname, operator_username, status_full = row
                # Дата и время
                date_fmt = ""; time_fmt = ""
                try:
                    dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
                    date_fmt = dt.strftime("%d.%m.%y")
                    time_fmt = dt.strftime("%H:%M:%S")
                except Exception:
                    try:
                        d = datetime.strptime(dt_str, "%Y-%m-%d")
                        date_fmt = d.strftime("%d.%m.%y")
                        time_fmt = ""
                    except Exception:
                        date_fmt = str(dt_str)
                        time_fmt = ""

                problem_text = str(problem).strip() if problem is not None else ""
                deadline_text = ""
                if problem_text:
                    m = re.search(r"\((\s*срок\s+выполнения\s*[^)]*)\)", problem_text, flags=re.IGNORECASE)
                    if m:
                        deadline_text = m.group(1).strip()
                        problem_text = re.sub(r"\((\s*срок\s+выполнения\s*[^)]*)\)", "", problem_text, flags=re.IGNORECASE).strip()

                applicant_line = " ".join([p for p in [applicant_name, applicant_surname] if p]).strip()

                formatted_rows.append({
                    "№ п/п": idx,
                    "Номер заявки": f"№ {rec_id}",
                    "Дата": date_fmt,
                    "Время": time_fmt,
                "Содержание заявки": problem_text.upper() if problem_text else "",
                    "Категория": category or "",
                    "Номер бригады": brigade_number or "",
                    "Срок выполнения": deadline_text,
                    "Телефон": phone or "",
                    "Адрес": address or "",
                    "Примечание": improvement or "",
                    "Заявитель": applicant_line,
                    "Оператор": operator_username or "",
                    "Состояние выполнения": status_full or "",
                })

            try:
                df = pd.DataFrame(formatted_rows, columns=[
                    "№ п/п", "Номер заявки", "Дата", "Время", "Содержание заявки", "Категория", "Номер бригады", "Срок выполнения", "Телефон", "Адрес", "Примечание", "Заявитель", "Оператор", "Состояние выполнения"
                ])
                default_name = "экспорт_текущий_список.xlsx"
                save_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    initialfile=default_name,
                    filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")],
                    initialdir=(export_base_dir_var.get() or "")
                )
                if not save_path:
                    messagebox.showinfo("Отмена", "Экспорт отменён пользователем")
                    return
                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Лист1")
                    ws = writer.sheets["Лист1"]
                    # Автоширины
                    for idx_c, col in enumerate(df.columns, start=1):
                        max_len = len(str(col))
                        for val in df[col].astype(str).values:
                            if val is None:
                                continue
                            for line in str(val).split("\n"):
                                if len(line) > max_len:
                                    max_len = len(line)
                        ws.column_dimensions[get_column_letter(idx_c)].width = min(80, max_len + 2)
                    # Фиксированные ширины
                    try:
                        name_to_width = {
                            "№ п/п": 8,
                            "Номер заявки": 16,
                            "Дата": 14,
                            "Время": 12,
                            "Содержание заявки": 50,
                            "Категория": 20,
                            "Номер бригады": 18,
                            "Срок выполнения": 28,
                            "Телефон": 20,
                            "Адрес": 40,
                            "Примечание": 30,
                            "Заявитель": 25,
                            "Оператор": 20,
                            "Состояние выполнения": 28,
                        }
                        for name, width in name_to_width.items():
                            if name in df.columns:
                                col_idx = list(df.columns).index(name) + 1
                                ws.column_dimensions[get_column_letter(col_idx)].width = width
                    except Exception:
                        pass

                    # Подписи "Составил/Утвердил" одна под другой
                    try:
                        sig_row = ws.max_row + 2
                        ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=2)
                        ws.cell(row=sig_row, column=1, value="Составил: __________________")
                        sig_row2 = sig_row + 1
                        ws.merge_cells(start_row=sig_row2, start_column=1, end_row=sig_row2, end_column=2)
                        ws.cell(row=sig_row2, column=1, value="Утвердил: __________________")
                    except Exception:
                        pass

                messagebox.showinfo("Экспорт", f"Файл успешно сохранён:\n{save_path}")
            except Exception as e:
                messagebox.showerror("Ошибка экспорта", f"Произошла ошибка при экспорте файла:\n{e}")

        def _export_full_period_local():
            """Экспорт без учёта локальных фильтров за период из этого окна (local_start/local_end)."""
            if not PANDAS_AVAILABLE:
                messagebox.showerror("Ошибка", "Модуль pandas не установлен.\n\n"
                                          "Установите зависимости:\n"
                                          "pip install pandas openpyxl\n\n"
                                          "Или запустите install_dependencies.bat")
                return
            try:
                from openpyxl.utils import get_column_letter
            except ImportError as e:
                messagebox.showerror("Ошибка", f"Не установлен модуль openpyxl: {e}\n\n"
                                          "pip install openpyxl")
                return

            # Выбор баз данных
            db_files = _select_multiple_databases()
            if db_files is None:
                return  # Пользователь отменил выбор

            try:
                start_date = local_start.get_date().strftime("%Y-%m-%d") if local_start.get() else None
                end_date = local_end.get_date().strftime("%Y-%m-%d") if local_end.get() else None
            except Exception:
                start_date = None
                end_date = None

            def execute_query(cursor_to_use):
                base_query = (
                    """SELECT r.id, r.problem, r.phone, r.address, r.improvement, r.brigade_number, r.category, COALESCE(r.created_at, r.date), r.name, r.surname, u.username, r.status
                                FROM records r 
                                LEFT JOIN common.users u ON r.user_id = u.id"""
                )
                params = []
                if start_date and end_date:
                    base_query += " WHERE date(r.date) >= date(?) AND date(r.date) <= date(?)"
                    params.extend([start_date, end_date])
                base_query += " ORDER BY r.id ASC"
                cursor_to_use.execute(base_query, params)
                return cursor_to_use.fetchall()
            
            try:
                rows = _execute_query_multiple_dbs(execute_query, db_files)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при получении данных из базы: {e}")
                return

            if not rows:
                messagebox.showinfo("Результат", "Записей в этом диапазоне нет")
                return

            formatted_rows = []
            for idx, row in enumerate(rows, start=1):
                rec_id, problem, phone, address, improvement, brigade_number, category, dt_str, applicant_name, applicant_surname, operator_username, status_full = row
                date_fmt = ""; time_fmt = ""
                try:
                    dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
                    date_fmt = dt.strftime("%d.%m.%y")
                    time_fmt = dt.strftime("%H:%M:%S")
                except Exception:
                    try:
                        d = datetime.strptime(dt_str, "%Y-%m-%d")
                        date_fmt = d.strftime("%d.%m.%y")
                        time_fmt = ""
                    except Exception:
                        date_fmt = str(dt_str)
                        time_fmt = ""

                problem_text = str(problem).strip() if problem is not None else ""
                deadline_text = ""
                if problem_text:
                    m = re.search(r"\((\s*срок\s+выполнения\s*[^)]*)\)", problem_text, flags=re.IGNORECASE)
                    if m:
                        deadline_text = m.group(1).strip()
                        problem_text = re.sub(r"\((\s*срок\s+выполнения\s*[^)]*)\)", "", problem_text, flags=re.IGNORECASE).strip()

                applicant_line = " ".join([p for p in [applicant_name, applicant_surname] if p]).strip()

                formatted_rows.append({
                    "№ п/п": idx,
                    "Номер заявки": f"№ {rec_id}",
                    "Дата": date_fmt or "",
                    "Время": time_fmt or "",
                    "Содержание заявки": problem_text.upper() if problem_text else "",
                    "Категория": category or "",
                    "Номер бригады": brigade_number or "",
                    "Срок выполнения": deadline_text or "",
                    "Телефон": phone or "",
                    "Адрес": address or "",
                    "Примечание": improvement or "",
                    "Заявитель": applicant_line or "",
                    "Оператор": operator_username or "",
                    "Состояние выполнения": status_full or "",
                })

            try:
                df = pd.DataFrame(formatted_rows, columns=[
                    "№ п/п", "Номер заявки", "Дата", "Время", "Содержание заявки", "Категория", "Номер бригады", "Срок выполнения", "Телефон", "Адрес", "Примечание", "Заявитель", "Оператор", "Состояние выполнения"
                ])
                default_name = f"отчёт_{local_start.get()}_to_{local_end.get()}.xlsx"
                save_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    initialfile=default_name,
                    filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")],
                    initialdir=(export_base_dir_var.get() or "")
                )
                if not save_path:
                    messagebox.showinfo("Отмена", "Экспорт отменён пользователем")
                    return
                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Лист1")
                    ws = writer.sheets["Лист1"]
                    
                    # Расширенные ширины колонок
                    try:
                        name_to_width = {
                            "№ п/п": 8,
                            "Номер заявки": 16,
                            "Дата": 14,
                            "Время": 12,
                            "Содержание заявки": 50,
                            "Категория": 20,
                            "Номер бригады": 18,
                            "Срок выполнения": 28,
                            "Телефон": 20,
                            "Адрес": 40,
                            "Примечание": 30,
                            "Заявитель": 25,
                            "Оператор": 20,
                            "Состояние выполнения": 28,
                        }
                        for name, width in name_to_width.items():
                            if name in df.columns:
                                col_idx = list(df.columns).index(name) + 1
                                ws.column_dimensions[get_column_letter(col_idx)].width = width
                    except Exception:
                        # Фолбэк: автоширины
                            for idx_c, col in enumerate(df.columns, start=1):
                                max_len = len(str(col))
                                for val in df[col].astype(str).values:
                                    if val is None:
                                        continue
                                    for line in str(val).split("\n"):
                                        if len(line) > max_len:
                                            max_len = len(line)
                                ws.column_dimensions[get_column_letter(idx_c)].width = min(80, max_len + 2)
                    
                    # Подписи "Составил/Утвердил"
                    try:
                        sig_row = ws.max_row + 2
                        ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=2)
                        ws.cell(row=sig_row, column=1, value="Составил: __________________")
                        sig_row2 = sig_row + 1
                        ws.merge_cells(start_row=sig_row2, start_column=1, end_row=sig_row2, end_column=2)
                        ws.cell(row=sig_row2, column=1, value="Утвердил: __________________")
                    except Exception:
                        pass
            except Exception as e:
                messagebox.showerror("Ошибка экспорта", f"Произошла ошибка при экспорте файла:\n{e}")

        def _export_filtered_local():
            """Экспорт по фильтрам в формат, как на образце: шапка с фильтрами и таблица."""
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Font, Border, Side
            except ImportError as e:
                messagebox.showerror("Ошибка", f"Не установлен openpyxl: {e}\n\npip install openpyxl")
                return

            # Выбор баз данных
            db_files = _select_multiple_databases()
            if db_files is None:
                return  # Пользователь отменил выбор

            # Собираем фильтры
            kw = (local_keyword.get().strip() or "")
            pv = (local_problem_var.get().strip() or "")
            sv = (local_status_var.get().strip() or "")
            ov = (local_operator_var.get().strip() or "")
            try:
                start_sql = local_start.get_date().strftime("%Y-%m-%d") if local_start.get() else None
                end_sql = local_end.get_date().strftime("%Y-%m-%d") if local_end.get() else None
                start_h = local_start.get_date().strftime("%d.%m.%y") if local_start.get() else ""
                end_h = local_end.get_date().strftime("%d.%m.%y") if local_end.get() else ""
            except Exception:
                start_sql = None; end_sql = None; start_h = ""; end_h = ""

            # Данные
            try:
                rows = _fetch_rows_for_local(kw or None, pv or None, sv or None, ov or None, start_sql, end_sql, db_files)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при получении данных: {e}")
                return
            if not rows:
                messagebox.showinfo("Результат", "Записей по текущим фильтрам нет")
                return

            wb = Workbook(); ws = wb.active; ws.title = "Отчёт"
            bold = Font(bold=True); thin = Side(style="thin"); border = Border(left=thin, right=thin, top=thin, bottom=thin)

            r = 1
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.cell(row=r, column=1, value="Сводка заявок").font = Font(bold=True, size=14)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
            r += 1
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.cell(row=r, column=1, value=f"за период с  {start_h or '...'}  по  {end_h or '...'} г.г.").alignment = Alignment(horizontal="center")
            r += 1

            parts = []
            if sv: parts.append(f"Состояние выполнения: {sv}")
            if pv: parts.append(f"Содержание: {pv}")
            if ov: parts.append(f"Оператор: {ov}")
            if kw: parts.append(f"Поиск: {kw}")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.cell(row=r, column=1, value=("Отбор: " + "; ".join(parts)) if parts else "Отбор: не задан").alignment = Alignment(horizontal="center")
            r += 2

            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.cell(row=r, column=1, value=f"Всего отобрано  {len(rows)} заявок.").font = bold
            r += 2

            headers = ["№ п/п", "Номер заявки", "Дата/Время обращения", "Содержание заявки", "Постановка на выполнение", "Состояние выполнения"]
            for c, h in enumerate(headers, start=1):
                ws.cell(row=r, column=c, value=h).font = bold
                ws.cell(row=r, column=c).border = border
                ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
            r += 1

            # Строки таблицы
            for i, row in enumerate(rows, start=1):
                # rows получены из _fetch_rows_for_local: там теперь есть r.category и r.status
                if len(row) >= 12:
                    rec_id, problem, phone, address, improvement, brigade_number, category, dt_str, applicant_name, applicant_surname, operator_username, status_full = row
                elif len(row) >= 11:
                    rec_id, problem, phone, address, improvement, brigade_number, category, dt_str, applicant_name, applicant_surname, operator_username = row
                    status_full = ""
                elif len(row) >= 10:
                    rec_id, problem, phone, address, improvement, dt_str, applicant_name, applicant_surname, operator_username, status_full = row
                    category = ""
                else:
                    rec_id, problem, phone, address, improvement, dt_str, applicant_name, applicant_surname, operator_username = row
                    status_full = ""
                    category = ""
                # Дата/время
                date_p = ""; time_p = ""
                try:
                    dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S"); date_p = dt.strftime("%d.%m.%y"); time_p = dt.strftime("%H:%M:%S")
                except Exception:
                    try:
                        d = datetime.strptime(dt_str, "%Y-%m-%d"); date_p = d.strftime("%d.%m.%y"); time_p = ""
                    except Exception:
                        date_p = str(dt_str or "")

                # Содержание и срок (часы)
                txt = str(problem or "").strip()
                hours = ""
                if txt:
                    m = re.search(r"\(\s*срок\s+выполнения\s*(\d+)\s*ч\s*\)", txt, flags=re.IGNORECASE)
                    if m:
                        hours = m.group(1)
                        txt = re.sub(r"\(\s*срок\s+выполнения\s*\d+\s*ч\s*\)", "", txt, flags=re.IGNORECASE).strip()

                fio = " ".join([p for p in [applicant_name, applicant_surname] if p]).strip()
                content_lines = []
                if category: content_lines.append(f"КАТЕГОРИЯ: {category.upper()}")
                if txt: content_lines.append(txt.upper())
                if phone: content_lines.append(f"ТЕЛ. {phone}")
                if address: content_lines.append(str(address))
                if fio: content_lines.append(fio)
                content_val = "\n".join(content_lines)

                ws.cell(row=r, column=1, value=i)
                ws.cell(row=r, column=2, value=f"№ {rec_id}")
                ws.cell(row=r, column=3, value=f"{date_p}\n{time_p}" if time_p else date_p)
                ws.cell(row=r, column=4, value=content_val)
                ws.cell(row=r, column=5, value=(f"СРОК ВЫПОЛНЕНИЯ: {hours} ч." if hours else ""))
                ws.cell(row=r, column=6, value=status_full)

                for c in range(1,7):
                    cell = ws.cell(row=r, column=c); cell.border = border
                    if c in (3,4,5):
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                    else:
                        cell.alignment = Alignment(vertical="top")
                r += 1

            # Ширины колонок
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 16
            ws.column_dimensions['D'].width = 58
            ws.column_dimensions['E'].width = 24
            ws.column_dimensions['F'].width = 26

            # Подпись
            sig = r + 1
            ws.merge_cells(start_row=sig, start_column=1, end_row=sig, end_column=3)
            ws.cell(row=sig, column=1, value="Отчёт сформирован:")

            default_name = f"сводка_по_фильтрам_{datetime.now().strftime('%d.%m.%Y_%H-%M')}.xlsx"
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_name, filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")], initialdir=(export_base_dir_var.get() or ""))
            if not save_path:
                messagebox.showinfo("Отмена", "Экспорт отменён пользователем"); return
            try:
                wb.save(save_path)
                messagebox.showinfo("Экспорт", f"Файл успешно сохранён:\n{save_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}")

        def _collect_summary_counts(start_date: str, end_date: str, db_files=None):
            """Собирает статистику по проблемам за период для сводных отчётов."""
            import re as _re

            WATER_PROBLEMS = [
                "течь воды",
                "течь в/к",
                "течь в / колонки",
                "течь из-под земли",
                "течь из-под асфальта",
                "течь трассы холодной воды",
                "течь пожарного гидранта",
                "дефект водоразборной колонки",
                "ржавая холодная вода в жилом фонде",
                "Не работает в / колонка",
                "слабое давление х/в",
                "нет х/в",
                "восстановить в/колонку",
                "откачка воды из колодца",
                "обвал в/к",
            ]
            SEWER_PROBLEMS = [
                "забой канализационного колодца",
                "течь канализации по дороге",
                "открыт колодец",
                "открыт колодец (отсутствие/несоответствие крышки)",
                "забит колодец",
                "обвал к/к",
                "привести к/к в нормативное состояние",
                "привести в / к в нормативное состояние",
                "разрушена плита колодца",
            ]

            def norm(p):
                p = _re.sub(r'\s*\(срок выполнения .+? ч\)\s*$', '', p or '', flags=_re.IGNORECASE).strip()
                p = p.replace('из -под', 'из-под')
                return ' '.join(p.split()).lower()

            lookup = {}
            for p in WATER_PROBLEMS + SEWER_PROBLEMS:
                lookup[norm(p)] = p

            def execute_query(cursor_to_use):
                cursor_to_use.execute(
                    """
                    SELECT problem, COUNT(*) as count
                    FROM records
                    WHERE date(date) >= date(?) AND date(date) <= date(?)
                    GROUP BY problem
                    """,
                    (start_date, end_date)
                )
                return cursor_to_use.fetchall()

            try:
                rows = _execute_query_multiple_dbs(execute_query, db_files)
                counts = {p: 0 for p in WATER_PROBLEMS + SEWER_PROBLEMS}

                for problem, count in rows:
                    canonical = lookup.get(norm(problem))
                    if canonical:
                        counts[canonical] += count

                water_total = sum(counts[p] for p in WATER_PROBLEMS)
                sewer_total = sum(counts[p] for p in SEWER_PROBLEMS)

                return counts, water_total, sewer_total, WATER_PROBLEMS, SEWER_PROBLEMS
            except Exception:
                all_p = WATER_PROBLEMS + SEWER_PROBLEMS
                return {p: 0 for p in all_p}, 0, 0, WATER_PROBLEMS, SEWER_PROBLEMS

        def _export_summary_local(period: str):
            """Краткие отчёты (сутки/неделя) по датам из этого окна."""
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment
            except ImportError as e:
                messagebox.showerror("Ошибка", f"Не установлен openpyxl: {e}\n\npip install openpyxl")
                return

            try:
                if period == "day":
                    start = local_start.get_date()
                    end = start
                    title_period = start.strftime("%d.%m.%Y")
                else:
                    end = local_end.get_date()
                    start = end - timedelta(days=6)
                    title_period = f"{start.strftime('%d.%m.%Y')} - {end.strftime('%d.%m.%Y')}"
                start_sql = start.strftime("%Y-%m-%d")
                end_sql = end.strftime("%Y-%m-%d")
                # Базы данных определяются автоматически по датам периода (не требуется ручной выбор)
                db_files = _db_files_for_date_range(start, end)
                counts, water_total, sewer_total, WATER_PROBLEMS, SEWER_PROBLEMS = _collect_summary_counts(start_sql, end_sql, db_files)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при подготовке данных: {e}")
                return

            default_name = (
                f"отчёт_за_сутки_{start.strftime('%d.%m.%Y')}.xlsx" if period == "day"
                else f"отчёт_за_неделю_{start.strftime('%d.%m.%Y')}_to_{end.strftime('%d.%m.%Y')}.xlsx"
            )
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")],
                initialdir=(export_base_dir_var.get() or ""),
            )
            if not save_path:
                messagebox.showinfo("Отмена", "Экспорт отменён пользователем")
                return

            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Отчёт"
                bold = Font(bold=True)
                row = 1
                ws.cell(row=row, column=1, value="Система водоснабжения").font = bold; row += 1
                ws.cell(row=row, column=1, value="Общее количество:").font = bold
                ws.cell(row=row, column=2, value=water_total); row += 1
                for p in WATER_PROBLEMS:
                    ws.cell(row=row, column=1, value=p)
                    ws.cell(row=row, column=2, value=counts[p])
                    row += 1
                row += 1
                ws.cell(row=row, column=1, value="Система водоотведения").font = bold; row += 1
                ws.cell(row=row, column=1, value="Общее кол-во:").font = bold
                ws.cell(row=row, column=2, value=sewer_total); row += 1
                for p in SEWER_PROBLEMS:
                    ws.cell(row=row, column=1, value=p)
                    ws.cell(row=row, column=2, value=counts[p])
                    row += 1
                row += 1
                ws.cell(row=row, column=1, value="Период").font = bold
                ws.cell(row=row, column=2, value=title_period); row += 1
                ws.cell(row=row, column=1, value="Смену сдал(а):"); ws.cell(row=row, column=2, value=""); row += 1
                ws.column_dimensions['A'].width = 50
                ws.column_dimensions['B'].width = 12
                try:
                    row += 2
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                    ws.cell(row=row, column=1, value="Составил: __________________")
                    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
                    row2 = row + 1
                    ws.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=3)
                    ws.cell(row=row2, column=1, value="Утвердил: __________________")
                    ws.cell(row=row2, column=1).alignment = Alignment(horizontal="left")
                except Exception:
                    pass
                wb.save(save_path)
                messagebox.showinfo("Экспорт", f"Отчёт успешно сохранён:\n{save_path}")
            except Exception as e:
                messagebox.showerror("Ошибка экспорта", f"Произошла ошибка при создании отчёта:\n{e}")

        def _export_daily_blank_local():
            """Бланк сведений за сутки по локальной дате начала."""
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, Border, Side
            except ImportError as e:
                messagebox.showerror("Ошибка", f"Не установлен openpyxl: {e}\n\npip install openpyxl")
                return

            try:
                day = local_start.get_date()
            except Exception:
                messagebox.showerror("Ошибка", "Выберите дату начала для формирования бланка")
                return
            day_sql = day.strftime("%Y-%m-%d")
            # База данных определяется автоматически по дате (не требуется ручной выбор)
            db_files = _db_files_for_date_range(day, day)

            def execute_query(cursor_to_use):
                cursor_to_use.execute(
                    """
                    SELECT r.problem, r.address, r.phone, r.name, r.surname, r.brigade_number, COALESCE(r.created_at, r.date), r.status, r.category
                    FROM records r
                    WHERE date(r.date) = date(?)
                    ORDER BY r.id ASC
                    """,
                    (day_sql,)
                )
                return cursor_to_use.fetchall()

            try:
                rows = _execute_query_multiple_dbs(execute_query, db_files)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка чтения из базы: {e}")
                return

            def category_for(problem_text: str) -> str:
                t = (problem_text or "").lower()
                
                # Сначала проверяем, есть ли категория в скобках в самом тексте проблемы
                # Формат: "содержание (категория)" - извлекаем категорию из скобок
                import re
                # Ищем категорию в скобках, но не "срок выполнения"
                category_match = re.search(r'\(([^)]+)\)', t)
                if category_match:
                    category_in_parens = category_match.group(1).lower().strip()
                    # Игнорируем "срок выполнения" и другие служебные скобки
                    if "срок" not in category_in_parens and "выполнен" not in category_in_parens:
                        # Маппинг категорий из скобок в названия секций бланка
                        category_map = {
                            "канализация": "КАНАЛИЗАЦИЯ",
                            "водопровод": "ВОДОПРОВОД",
                            "водоотведение": "КАНАЛИЗАЦИЯ",
                            "водоснабжение": "ВОДОПРОВОД",
                            "перекладка": "ПЕРЕКЛАДКА",
                            "в/колонки": "В/КОЛОНКИ",
                            "колонки": "В/КОЛОНКИ",
                            "пожарные гидранты": "ПОЖАРНЫЕ ГИДРАНТЫ",
                            "гидранты": "ПОЖАРНЫЕ ГИДРАНТЫ",
                            "частные врезки": "ЧАСТНЫЕ ВРЕЗКИ",
                            "врезки": "ЧАСТНЫЕ ВРЕЗКИ",
                        }
                        # Проверяем точное совпадение или частичное
                        for key, value in category_map.items():
                            if key in category_in_parens:
                                return value
                            # Если не нашли точное совпадение, пробуем по первым словам
                            if "канализац" in category_in_parens:
                                return "КАНАЛИЗАЦИЯ"
                            if "водопровод" in category_in_parens or "водоснабж" in category_in_parens:
                                return "ВОДОПРОВОД"
                            if "перекладк" in category_in_parens:
                                return "ПЕРЕКЛАДКА"
                            if "колонк" in category_in_parens:
                                return "В/КОЛОНКИ"
                            if "гидрант" in category_in_parens:
                                return "ПОЖАРНЫЕ ГИДРАНТЫ"
                            if "врезк" in category_in_parens:
                                return "ЧАСТНЫЕ ВРЕЗКИ"
                
                # Проверяем маппинг из problem_to_blank_category
                problem_clean = clean_problem_text(problem_text).lower()
                if problem_clean in problem_to_blank_category:
                    return problem_to_blank_category[problem_clean]
                
                # ПЕРЕКЛАДКА: прокладка водопровода, врезка водопровода, разрытие, прокол канализации, прокол водопровода, замена водовода
                # Приоритет: проверяем сначала ПЕРЕКЛАДКУ
                if "перекладк" in t or "прокладк" in t:
                    # Если есть "перекладка" или "прокладка" - это ПЕРЕКЛАДКА
                    return "ПЕРЕКЛАДКА"
                if "разрытие" in t and "восстанов" not in t:
                    # "разрытие" само по себе - ПЕРЕКЛАДКА, но "восстановить разрытие" - ВОДОПРОВОД
                    return "ПЕРЕКЛАДКА"
                if "прокол" in t:
                    # прокол канализации или прокол водопровода
                    return "ПЕРЕКЛАДКА"
                if "замен" in t and "водовод" in t:
                    # замена водовода
                    return "ПЕРЕКЛАДКА"
                if "врезк" in t and "водопровод" in t and "частн" not in t:
                    # врезка водопровода (но не частная)
                    return "ПЕРЕКЛАДКА"
                
                # ЧАСТНЫЕ ВРЕЗКИ: ч/врезка, ч/врезка (нужна откачка воды из колодца)
                if ("врезк" in t and "частн" in t) or ("ч/" in t and "врезк" in t) or \
                   ("частн" in t and "врезк" in t):
                    return "ЧАСТНЫЕ ВРЕЗКИ"
                
                # ПОЖАРНЫЕ ГИДРАНТЫ: неисправен ПГ, домонтирован ПГ, нет воды в ПГ, соран шток ПГ и затоплен водой
                if "гидрант" in t or "пг" in t:
                    return "ПОЖАРНЫЕ ГИДРАНТЫ"
                
                # В/КОЛОНКИ: ремонт в/колонки
                if "колонк" in t:
                    return "В/КОЛОНКИ"
                
                # ВОДОПРОВОД: перекладка водопровода, течь, замена крана (оплачено), течь, открыт в/к и течь в/к, 
                # перекрыть/открыть х/в (оплачено), восстановить благоустройство, сл.давление х/воды, восстановить разрытие
                if "течь" in t and "канализац" not in t:
                    # течь (но не течь канализации)
                    return "ВОДОПРОВОД"
                if "замен" in t and "кран" in t:
                    # замена крана
                    return "ВОДОПРОВОД"
                if "открыт" in t and ("в/к" in t or "в/колонк" in t):
                    # открыт в/к и течь в/к
                    return "ВОДОПРОВОД"
                if ("перекрыт" in t or "открыт" in t) and ("х/в" in t or "холодн" in t):
                    # перекрыть/открыть х/в
                    return "ВОДОПРОВОД"
                if "восстанов" in t and ("благоустр" in t or "разрытие" in t):
                    # восстановить благоустройство, восстановить разрытие
                    return "ВОДОПРОВОД"
                if ("слаб" in t or "сл." in t) and ("давл" in t or "х/в" in t or "холодн" in t):
                    # сл.давление х/воды
                    return "ВОДОПРОВОД"
                if "водопровод" in t:
                    # все остальное с водопроводом
                    return "ВОДОПРОВОД"
                
                # КАНАЛИЗАЦИЯ: все остальное, связанное с канализацией
                if ("засор" in t) or ("колодец" in t and ("канализац" in t or "забой" in t)) or ("канализац" in t):
                    return "КАНАЛИЗАЦИЯ"
                
                # По умолчанию - ВОДОПРОВОД
                return "ВОДОПРОВОД"

            sections = {
                "ПЕРЕКЛАДКА": [],
                "ВОДОПРОВОД": [],
                "В/КОЛОНКИ": [],
                "ПОЖАРНЫЕ ГИДРАНТЫ": [],
                "ЧАСТНЫЕ ВРЕЗКИ": [],
                "КАНАЛИЗАЦИЯ": [],
            }

            for rec in rows:
                if len(rec) >= 9:
                    problem, address, phone, name_, surname_, brigade_number, dt_str, status_full, category = rec
                else:
                    problem, address, phone, name_, surname_, brigade_number, dt_str, status_full = rec
                    category = ""
                who = " ".join([p for p in [name_ or "", surname_ or ""] if p]).strip()
                time_part = ""; date_part = ""
                try:
                    from datetime import datetime as _dt
                    dt = _dt.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
                    time_part = dt.strftime("%H:%M"); date_part = dt.strftime("%d.%m.%y")
                except Exception:
                    time_part = ""; date_part = ""
                line_parts = []
                if time_part:
                    line_parts.append(time_part)
                if category:
                    line_parts.append(f"Категория: {category}")
                if address:
                    line_parts.append(str(address))
                if problem:
                    line_parts.append(str(problem))
                if phone:
                    line_parts.append(f"тел: {phone}")
                if who:
                    line_parts.append(who)
                if brigade_number:
                    # Отображаем номер бригады без суффикса .бр если он есть
                    brigade_display = str(brigade_number).replace(".бр", "").strip()
                    if brigade_display:
                        line_parts.append(f"бригада: {brigade_display}")
                text_line = " — ".join(line_parts)
                # Правый текст: дата/время и статус
                right_text = ""
                try:
                    right_text = (date_part + (" " + time_part if time_part else "")).strip()
                    if status_full:
                        right_text = (right_text + (" — " if right_text else "") + status_full).strip()
                except Exception:
                    right_text = status_full or ""
                sections[category_for(problem)].append((text_line, right_text))

            wb = Workbook()
            ws = wb.active
            ws.title = "Бланк"
            title = f"БЛАНК - СВЕДЕНИЙ  {day.strftime('%d.%m.%y')}"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
            ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

            thin = Side(style="thin")
            border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

            row_idx = 3
            for section_name, lines in sections.items():
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
                ws.cell(row=row_idx, column=1, value=section_name).font = Font(bold=True)
                ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")
                row_idx += 1
                start_box = row_idx
                if not lines:
                    lines = [("", "")]
                for line in lines:
                    # line может быть строкой из старых форматов — приведём к кортежу
                    if isinstance(line, tuple):
                        left_text, right_text = line
                    else:
                        left_text, right_text = (str(line), "")
                    # Слева объединяем столбцы A..E, справа пишем статус/дату в F с выравниванием вправо
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
                    ws.cell(row=row_idx, column=1, value=left_text)
                    ws.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True)
                    ws.cell(row=row_idx, column=6, value=right_text)
                    ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right")
                    row_idx += 1
                for r in range(start_box - 1, row_idx):
                    for c in range(1, 7):
                        ws.cell(row=r, column=c).border = border_all
                row_idx += 1

            ws.column_dimensions['A'].width = 16
            for col in ['B', 'C', 'D', 'E']:
                ws.column_dimensions[col].width = 24
            ws.column_dimensions['F'].width = 44
            try:
                sig_row = ws.max_row + 2
                ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=3)
                ws.cell(row=sig_row, column=1, value="Составил: __________________")
                sig_row2 = sig_row + 1
                ws.merge_cells(start_row=sig_row2, start_column=1, end_row=sig_row2, end_column=3)
                ws.cell(row=sig_row2, column=1, value="Утвердил: __________________")
            except Exception:
                pass

            default_name = f"бланк_сведений_{day.strftime('%d.%m.%Y')}.xlsx"
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")],
                initialdir=(export_base_dir_var.get() or ""),
            )
            if not save_path:
                messagebox.showinfo("Отмена", "Экспорт отменён пользователем")
                return
            try:
                wb.save(save_path)
                messagebox.showinfo("Экспорт", f"Бланк сохранён:\n{save_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}")

        def _export_daily_blank_custom():
            """Бланк сведений за выбранную дату."""
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, Border, Side
                from tkcalendar import DateEntry
            except ImportError as e:
                messagebox.showerror("Ошибка", f"Не установлен openpyxl или tkcalendar: {e}\n\npip install openpyxl tkcalendar")
                return

            # Окно выбора даты
            date_win = tk.Toplevel(main)
            setup_scaling()
            date_win.title("Выбор даты для бланка")
            date_win.resizable(False, False)
            try:
                fit_and_center_window(date_win, min_width=300, min_height=120)
            except:
                pass

            frm = ttk.Frame(date_win, padding="20")
            frm.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            date_win.columnconfigure(0, weight=1)
            date_win.rowconfigure(0, weight=1)

            ttk.Label(frm, text="Выберите дату:").grid(row=0, column=0, sticky=tk.W, pady=5)
            date_entry = DateEntry(frm, date_pattern='dd.MM.yyyy', width=15)
            date_entry.grid(row=0, column=1, sticky=tk.W, padx=(10, 0), pady=5)
            date_entry.set_date(datetime.now())

            def export_with_date():
                try:
                    day = date_entry.get_date()
                except Exception:
                    messagebox.showerror("Ошибка", "Выберите корректную дату")
                    return
                date_win.destroy()
                
                # Выбор баз данных
                db_files = _select_multiple_databases()
                if db_files is None:
                    return  # Пользователь отменил выбор
                
                day_sql = day.strftime("%Y-%m-%d")

                def execute_query(cursor_to_use):
                    cursor_to_use.execute(
                        """
                        SELECT r.problem, r.address, r.phone, r.name, r.surname, r.brigade_number, COALESCE(r.created_at, r.date), r.status, r.category
                        FROM records r
                        WHERE date(r.date) = date(?)
                        ORDER BY r.id ASC
                        """,
                        (day_sql,)
                    )
                    return cursor_to_use.fetchall()

                try:
                    rows = _execute_query_multiple_dbs(execute_query, db_files)
                except Exception as e:
                    messagebox.showerror("Ошибка", f"Ошибка чтения из базы: {e}")
                    return

                def category_for(problem_text: str) -> str:
                    t = (problem_text or "").lower()
                    
                    # Сначала проверяем, есть ли категория в скобках в самом тексте проблемы
                    # Формат: "содержание (категория)" - извлекаем категорию из скобок
                    import re
                    # Ищем категорию в скобках, но не "срок выполнения"
                    category_match = re.search(r'\(([^)]+)\)', t)
                    if category_match:
                        category_in_parens = category_match.group(1).lower().strip()
                        # Игнорируем "срок выполнения" и другие служебные скобки
                        if "срок" not in category_in_parens and "выполнен" not in category_in_parens:
                            # Маппинг категорий из скобок в названия секций бланка
                            category_map = {
                                "канализация": "КАНАЛИЗАЦИЯ",
                                "водопровод": "ВОДОПРОВОД",
                                "водоотведение": "КАНАЛИЗАЦИЯ",
                                "водоснабжение": "ВОДОПРОВОД",
                                "перекладка": "ПЕРЕКЛАДКА",
                                "в/колонки": "В/КОЛОНКИ",
                                "колонки": "В/КОЛОНКИ",
                                "пожарные гидранты": "ПОЖАРНЫЕ ГИДРАНТЫ",
                                "гидранты": "ПОЖАРНЫЕ ГИДРАНТЫ",
                                "частные врезки": "ЧАСТНЫЕ ВРЕЗКИ",
                                "врезки": "ЧАСТНЫЕ ВРЕЗКИ",
                            }
                            # Проверяем точное совпадение или частичное
                            for key, value in category_map.items():
                                if key in category_in_parens:
                                    return value
                            # Если не нашли точное совпадение, пробуем по первым словам
                            if "канализац" in category_in_parens:
                                return "КАНАЛИЗАЦИЯ"
                            if "водопровод" in category_in_parens or "водоснабж" in category_in_parens:
                                return "ВОДОПРОВОД"
                            if "перекладк" in category_in_parens:
                                return "ПЕРЕКЛАДКА"
                            if "колонк" in category_in_parens:
                                return "В/КОЛОНКИ"
                            if "гидрант" in category_in_parens:
                                return "ПОЖАРНЫЕ ГИДРАНТЫ"
                            if "врезк" in category_in_parens:
                                return "ЧАСТНЫЕ ВРЕЗКИ"
                    
                    # Проверяем маппинг из problem_to_blank_category (используем оригинальный problem_text)
                    try:
                        problem_clean = clean_problem_text(problem_text).lower()
                        if problem_clean in problem_to_blank_category:
                            return problem_to_blank_category[problem_clean]
                    except Exception:
                        pass
                    
                    if "перекладк" in t or "прокладк" in t:
                        return "ПЕРЕКЛАДКА"
                    if "разрытие" in t and "восстанов" not in t:
                        return "ПЕРЕКЛАДКА"
                    if "прокол" in t:
                        return "ПЕРЕКЛАДКА"
                    if "замен" in t and "водовод" in t:
                        return "ПЕРЕКЛАДКА"
                    if "врезк" in t and "водопровод" in t and "частн" not in t:
                        return "ПЕРЕКЛАДКА"
                    
                    if ("врезк" in t and "частн" in t) or ("ч/" in t and "врезк" in t) or \
                       ("частн" in t and "врезк" in t):
                        return "ЧАСТНЫЕ ВРЕЗКИ"
                    
                    if "гидрант" in t or "пг" in t:
                        return "ПОЖАРНЫЕ ГИДРАНТЫ"
                    
                    if "колонк" in t:
                        return "В/КОЛОНКИ"
                    
                    if "течь" in t and "канализац" not in t:
                        return "ВОДОПРОВОД"
                    if "замен" in t and "кран" in t:
                        return "ВОДОПРОВОД"
                    if "открыт" in t and ("в/к" in t or "в/колонк" in t):
                        return "ВОДОПРОВОД"
                    if ("перекрыт" in t or "открыт" in t) and ("х/в" in t or "холодн" in t):
                        return "ВОДОПРОВОД"
                    if "восстанов" in t and ("благоустр" in t or "разрытие" in t):
                        return "ВОДОПРОВОД"
                    if ("слаб" in t or "сл." in t) and ("давл" in t or "х/в" in t or "холодн" in t):
                        return "ВОДОПРОВОД"
                    if "водопровод" in t:
                        return "ВОДОПРОВОД"
                    
                    if ("засор" in t) or ("колодец" in t and ("канализац" in t or "забой" in t)) or ("канализац" in t):
                        return "КАНАЛИЗАЦИЯ"
                    
                    return "ВОДОПРОВОД"

                sections = {
                    "ПЕРЕКЛАДКА": [],
                    "ВОДОПРОВОД": [],
                    "В/КОЛОНКИ": [],
                    "ПОЖАРНЫЕ ГИДРАНТЫ": [],
                    "ЧАСТНЫЕ ВРЕЗКИ": [],
                    "КАНАЛИЗАЦИЯ": [],
                }

                for rec in rows:
                    if len(rec) >= 9:
                        problem, address, phone, name_, surname_, brigade_number, dt_str, status_full, category = rec
                    else:
                        problem, address, phone, name_, surname_, brigade_number, dt_str, status_full = rec
                        category = ""
                    who = " ".join([p for p in [name_ or "", surname_ or ""] if p]).strip()
                    time_part = ""; date_part = ""
                    try:
                        from datetime import datetime as _dt
                        dt = _dt.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
                        time_part = dt.strftime("%H:%M"); date_part = dt.strftime("%d.%m.%y")
                    except Exception:
                        time_part = ""; date_part = ""
                    line_parts = []
                    if time_part:
                        line_parts.append(time_part)
                    if category:
                        line_parts.append(f"Категория: {category}")
                    if address:
                        line_parts.append(str(address))
                    if problem:
                        line_parts.append(str(problem))
                    if phone:
                        line_parts.append(f"тел: {phone}")
                    if who:
                        line_parts.append(who)
                    if brigade_number:
                        # Отображаем номер бригады без суффикса .бр если он есть
                        brigade_display = str(brigade_number).replace(".бр", "").strip()
                        if brigade_display:
                            line_parts.append(f"бригада: {brigade_display}")
                    text_line = " — ".join(line_parts)
                    right_text = ""
                    try:
                        right_text = (date_part + (" " + time_part if time_part else "")).strip()
                        if status_full:
                            right_text = (right_text + (" — " if right_text else "") + status_full).strip()
                    except Exception:
                        right_text = status_full or ""
                    sections[category_for(problem)].append((text_line, right_text))

                wb = Workbook()
                ws = wb.active
                ws.title = "Бланк"
                title = f"БЛАНК - СВЕДЕНИЙ  {day.strftime('%d.%m.%y')}"
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
                ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
                ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

                thin = Side(style="thin")
                border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

                row_idx = 3
                for section_name, lines in sections.items():
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
                    ws.cell(row=row_idx, column=1, value=section_name).font = Font(bold=True)
                    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")
                    row_idx += 1
                    start_box = row_idx
                    if not lines:
                        lines = [("", "")]
                    for line in lines:
                        if isinstance(line, tuple):
                            left_text, right_text = line
                        else:
                            left_text, right_text = (str(line), "")
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
                        ws.cell(row=row_idx, column=1, value=left_text)
                        ws.cell(row=row_idx, column=1).alignment = Alignment(wrap_text=True)
                        ws.cell(row=row_idx, column=6, value=right_text)
                        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right")
                        row_idx += 1
                    for r in range(start_box - 1, row_idx):
                        for c in range(1, 7):
                            ws.cell(row=r, column=c).border = border_all
                    row_idx += 1

                ws.column_dimensions['A'].width = 16
                for col in ['B', 'C', 'D', 'E']:
                    ws.column_dimensions[col].width = 24
                ws.column_dimensions['F'].width = 44
                try:
                    sig_row = ws.max_row + 2
                    ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=3)
                    ws.cell(row=sig_row, column=1, value="Составил: __________________")
                    sig_row2 = sig_row + 1
                    ws.merge_cells(start_row=sig_row2, start_column=1, end_row=sig_row2, end_column=3)
                    ws.cell(row=sig_row2, column=1, value="Утвердил: __________________")
                except Exception:
                    pass

                default_name = f"бланк_сведений_{day.strftime('%d.%m.%Y')}.xlsx"
                save_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    initialfile=default_name,
                    filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")],
                    initialdir=(export_base_dir_var.get() or ""),
                )
                if not save_path:
                    messagebox.showinfo("Отмена", "Экспорт отменён пользователем")
                    return
                try:
                    wb.save(save_path)
                    messagebox.showinfo("Экспорт", f"Бланк сохранён:\n{save_path}")
                except Exception as e:
                    messagebox.showerror("Ошибка", f"Ошибка сохранения: {e}")

            btns = ttk.Frame(frm, padding="10")
            btns.grid(row=1, column=0, columnspan=2, sticky=(tk.E, tk.W))
            ttk.Button(btns, text="Сформировать", command=export_with_date).grid(row=0, column=0, padx=5)
            ttk.Button(btns, text="Отмена", command=date_win.destroy).grid(row=0, column=1, padx=5)

        # Кнопка с выпадающим меню
        try:
            export_menu = tk.Menu(filters_panel, tearoff=0)
            export_menu.add_command(label="Экспорт текущего списка (локальные фильтры)", command=_export_current_list)
            export_menu.add_command(label="Экспорт за период (локальные даты)", command=_export_full_period_local)
            export_menu.add_command(label="Экспорт с фильтрами (локальные)", command=_export_filtered_local)
            export_menu.add_separator()
            export_menu.add_command(label="Отчёт за сутки (локальная дата начала)", command=lambda: _export_summary_local("day"))
            export_menu.add_command(label="Отчёт за неделю (локальная дата конца)", command=lambda: _export_summary_local("week"))
            export_menu.add_command(label="Бланк сведений за сутки (локальная дата начала)", command=_export_daily_blank_local)
            export_menu.add_command(label="Бланк сведений (выбор даты)", command=_export_daily_blank_custom)

            def open_export_menu(event=None):
                try:
                    export_menu.tk_popup(btns_local.winfo_rootx() + btns_local.winfo_width(), btns_local.winfo_rooty())
                finally:
                    try:
                        export_menu.grab_release()
                    except Exception:
                        pass

            exp_btn = ttk.Button(btns_local, text="Экспорт ▾", command=open_export_menu)
            exp_btn.grid(row=0, column=2, padx=(6, 0))
        except Exception:
            pass

        frame_list = ttk.LabelFrame(records_window, text="Список данных", padding="12")
        frame_list.grid(row=2, column=0, padx=12, pady=(4, 8), sticky="nsew")
        records_window.columnconfigure(0, weight=1)
        records_window.rowconfigure(2, weight=1)
        frame_list.columnconfigure(0, weight=1)
        frame_list.rowconfigure(0, weight=1)

        columns = ('ID', 'Наименование', 'Фамилия', 'Категория', 'Содержание заявки', 'Номер бригады', 'Телефон', 'Адрес', 'Дата', 'Время', 'Постановка', 'Состояние', 'Пользователь')
        records_tree = ttk.Treeview(frame_list, columns=columns, show='headings', height=22)
        for col in columns:
            records_tree.heading(col, text=col)
            if col == 'ID':
                records_tree.column(col, width=70, minwidth=60, stretch=False)
            elif col in ('Наименование', 'Фамилия', 'Пользователь'):
                records_tree.column(col, width=130, minwidth=100, stretch=False)
            elif col == 'Категория':
                records_tree.column(col, width=120, minwidth=100, stretch=False)
            elif col == 'Номер бригады':
                records_tree.column(col, width=120, minwidth=100, stretch=False)
            elif col == 'Телефон':
                records_tree.column(col, width=140, minwidth=120, stretch=False)
            elif col in ('Дата', 'Время', 'Постановка', 'Состояние'):
                records_tree.column(col, width=130, minwidth=110, stretch=False)
            elif col == 'Адрес':
                records_tree.column(col, width=320, minwidth=220, stretch=True)
            elif col == 'Содержание заявки':
                records_tree.column(col, width=380, minwidth=260, stretch=True)
        records_tree.tag_configure('odd', background='#f9fafb')
        records_tree.tag_configure('even', background='#ffffff')
        records_tree.tag_configure('header', background='#e0e0e0', font=('', 10, 'bold'))
        scrollbar_y = ttk.Scrollbar(frame_list, orient="vertical", command=records_tree.yview)
        scrollbar_x = ttk.Scrollbar(frame_list, orient="horizontal", command=records_tree.xview)
        records_tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        records_tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")

        # Копирование выбранной строки в буфер обмена (двойной клик по строке или Ctrl+C)
        def _copy_selected_row(event=None):
            try:
                sel = records_tree.selection()
                if not sel:
                    return
                item_id = sel[0]
                values = records_tree.item(item_id).get('values', [])
                # Используем заголовки столбцов для подписей
                try:
                    col_names = list(columns)
                except Exception:
                    col_names = []
                lines = []
                for idx, val in enumerate(values):
                    try:
                        label = col_names[idx] if idx < len(col_names) else f"Колонка {idx+1}"
                    except Exception:
                        label = f"Колонка {idx+1}"
                    lines.append(f"{label}: {val if val is not None else ''}")
                text_to_copy = "\n".join(lines) if lines else ""
                if not text_to_copy:
                    return
                try:
                    records_window.clipboard_clear()
                    records_window.clipboard_append(text_to_copy)
                except Exception:
                    try:
                        main.clipboard_clear()
                        main.clipboard_append(text_to_copy)
                    except Exception:
                        return
                try:
                    show_auto_close_info("Скопировано в буфер обмена", duration_ms=2500)
                except Exception:
                    pass
            except Exception:
                pass

        try:
            records_tree.bind("<Double-1>", _copy_selected_row)
        except Exception:
            pass
        try:
            records_window.bind("<Control-c>", lambda e: _copy_selected_row())
        except Exception:
            pass

        # ---- Кнопки администратора ----
        def delete_record():
            selection = records_tree.selection()
            if not selection:
                messagebox.showwarning("Ошибка", "Выберите запись")
                return
            record_id = records_tree.item(selection[0])['values'][0]

            # Проверка прав: обычный пользователь может удалять только свои записи
            if user[3] != "admin":
                cursor.execute("SELECT user_id FROM records WHERE id=?", (record_id,))
                rec = cursor.fetchone()
                if not rec or rec[0] != user[0]:
                    messagebox.showwarning("Ошибка", "Вы можете удалять только свои записи")
                    return

            confirm = messagebox.askyesno("Подтверждение", "Удалить выбранную запись?")
            if not confirm:
                return

            cursor.execute("DELETE FROM records WHERE id=?", (record_id,))
            conn.commit()
            refresh_records_default()
            try:
                refresh_recent()
            except Exception:
                pass

        def edit_record():
            selection = records_tree.selection()
            if not selection:
                messagebox.showwarning("Ошибка", "Выберите запись")
                return
            record_id = records_tree.item(selection[0])['values'][0]

            # Все пользователи могут редактировать любые записи
            cursor.execute("SELECT name, surname, problem, phone, address, assignment_date, status, improvement, brigade_number, category FROM records WHERE id= ?", (record_id,))
            rec = cursor.fetchone()
            if not rec:
                return

            edit_win = tk.Toplevel(main)
            setup_scaling()
            edit_win.title("Редактировать запись")
            edit_win.resizable(True, True)
            try:
                fit_and_center_window(edit_win, min_width=520, min_height=550)
            except:
                pass

            frm = ttk.Frame(edit_win, padding="20")
            frm.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            edit_win.columnconfigure(0, weight=1)
            edit_win.rowconfigure(0, weight=1)
            frm.columnconfigure(1, weight=1)

            ttk.Label(frm, text="Наименование:").grid(row=0, column=0, sticky=tk.W, pady=5)
            name_entry = ttk.Entry(frm, width=30)
            name_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            name_entry.insert(0, rec[0])

            ttk.Label(frm, text="Фамилия:").grid(row=1, column=0, sticky=tk.W, pady=5)
            surname_entry = ttk.Entry(frm, width=30)
            surname_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            surname_entry.insert(0, rec[1])

            # Определение категорий (такие же как в основной форме)
            problem_categories_edit = {
                "водоотведение": [
                    "забой канализационного колодца",
                    "течь канализации по дороге",
                ],
                "водоснабжение": [
                    "течь воды",
                    "течь в / колонки",
                    "течь из-под земли",
                    "течь пожарного гидранта",
                    "течь из -под асфальта",
                    "течь трассы холодной воды",
                    "откачка воды из колодца",
                    "дефект водоразборной колонки",
                    "ржавая холодная вода в жилом фонде",
                    "Не работает в / колонка",
                    "слабое давление х/в",
                    "восстановить в/колонку",
                    "течь в/к",
                    "нет х/в"
                ],
                "общие": [
                    "перекладка",
                    "водопровод",
                    "частные врезки",
                    "открыт колодец (отсутствие/несоответствие крышки)",
                    "восстановление асфальтобетонного покрытия",
                    "разрушена плита колодца",
                    "привести к/к в нормативное состояние",
                    "привести в / к в нормативное состояние",
                    "обвал в/к",
                    "обвал к/к",
                ]
            }
            
            # Загружаем категорию из записи
            saved_category = rec[9] if len(rec) > 9 else ""
            category_display = saved_category if saved_category else ""
            
            ttk.Label(frm, text="Категория:").grid(row=2, column=0, sticky=tk.W, pady=5)
            category_var_edit = tk.StringVar(value=category_display)
            category_combo_edit = ttk.Combobox(frm, textvariable=category_var_edit, values=["", "Водоотведение", "Водоснабжение"], state="readonly", width=27)
            category_combo_edit.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

            ttk.Label(frm, text="Содержание заявки:").grid(row=3, column=0, sticky=tk.W, pady=5)
            # Разбираем исходный текст проблемы: отделяем срок выполнения, если он был сохранён ранее
            _orig_problem_text = rec[2] or ""
            try:
                # Ищем число часов в самых распространённых форматах
                _m_deadline = (
                    re.search(r"\(\s*срок\s+выполнения\s*(\d+)\s*ч\s*\)", _orig_problem_text, flags=re.IGNORECASE)
                    or re.search(r"срок\s+выполнения\s*[:\-]?\s*(\d+)\s*ч", _orig_problem_text, flags=re.IGNORECASE)
                    or re.search(r"\b(\d+)\s*ч\b", _orig_problem_text)
                )
                _prefill_deadline = _m_deadline.group(1) if _m_deadline else ""
                # Удаляем фрагмент со сроком из текста проблемы, если он в скобках
                _base_problem_text = re.sub(r"\(\s*срок\s+выполнения\s*\d+\s*ч\s*\)", "", _orig_problem_text, flags=re.IGNORECASE).strip()
            except Exception:
                _prefill_deadline = ""
                _base_problem_text = _orig_problem_text
            problem_var = tk.StringVar(value=_base_problem_text)
            
            def update_problem_options_edit(*args):
                """Обновляет список проблем в зависимости от выбранной категории."""
                selected_category = category_var_edit.get().lower()
                if selected_category == "водоотведение":
                    filtered_options = problem_categories_edit["водоотведение"]
                elif selected_category == "водоснабжение":
                    filtered_options = problem_categories_edit["водоснабжение"]
                else:
                    filtered_options = problem_categories_edit["общие"]
                
                problem_entry_edit['values'] = sorted(filtered_options, key=lambda s: s.lower())
                # Если текущее значение не входит в новый список, оставляем его (для редактирования существующих записей)
            
            category_var_edit.trace('w', update_problem_options_edit)
            
            # Определяем начальный список проблем на основе сохраненной категории
            if category_display.lower() == "водоотведение":
                initial_options = problem_categories_edit["водоотведение"]
            elif category_display.lower() == "водоснабжение":
                initial_options = problem_categories_edit["водоснабжение"]
            else:
                initial_options = problem_categories_edit["общие"]
            
            problem_entry_edit = ttk.Combobox(frm, textvariable=problem_var, values=sorted(initial_options, key=lambda s: s.lower()), state="readonly", width=27)
            problem_entry_edit.grid(row=3, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

            ttk.Label(frm, text="Номер бригады:").grid(row=4, column=0, sticky=tk.W, pady=5)
            brigade_var = tk.StringVar()
            brigade_entry = ttk.Entry(frm, textvariable=brigade_var, width=27)
            brigade_entry.grid(row=4, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            # Извлекаем номер бригады без суффикса .бр для отображения
            brigade_display = (rec[8] or "").replace(".бр", "") if rec[8] else ""
            brigade_entry.insert(0, brigade_display)

            ttk.Label(frm, text="Срок выполнения (часы):").grid(row=5, column=0, sticky=tk.W, pady=5)
            deadline_var = tk.StringVar(value=_prefill_deadline)
            deadline_entry = ttk.Entry(frm, textvariable=deadline_var, width=27)
            deadline_entry.grid(row=5, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

            ttk.Label(frm, text="Телефон звонившего:").grid(row=6, column=0, sticky=tk.W, pady=5)
            phone_entry = ttk.Entry(frm, width=30)
            phone_entry.grid(row=6, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            phone_entry.insert(0, rec[3] or "")

            ttk.Label(frm, text="Адрес:").grid(row=7, column=0, sticky=tk.W, pady=5)
            address_entry = ttk.Entry(frm, width=30)
            address_entry.grid(row=7, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            address_entry.insert(0, rec[4] or "")

            ttk.Label(frm, text="Постановка на выполнение:").grid(row=8, column=0, sticky=tk.W, pady=5)
            assignment_entry = DateEntry(frm, date_pattern='dd.MM.yyyy', width=22)
            assignment_entry.grid(row=8, column=1, sticky=tk.W, padx=(10, 0), pady=5)
            if rec[5]:
                try:
                    assignment_entry.set_date(datetime.strptime(rec[5], "%Y-%m-%d"))
                except:
                    pass

            ttk.Label(frm, text="Состояние выполнения:").grid(row=9, column=0, sticky=tk.W, pady=5)
            try:
                _status_base = re.sub(r"\s*\(\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2}\)\s*$", "", (rec[6] or "не выполнено")).strip()
            except Exception:
                _status_base = (rec[6] or "не выполнено")
            status_var = tk.StringVar(value=_status_base)
            status_combo = ttk.Combobox(frm, textvariable=status_var, values=["в работе", "не выполнено", "выполнено"], state="readonly", width=27)
            status_combo.grid(row=9, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)

            ttk.Label(frm, text="Примечание:").grid(row=10, column=0, sticky=tk.W, pady=5)
            improvement_entry = ttk.Entry(frm, width=30)
            improvement_entry.grid(row=10, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=5)
            improvement_entry.insert(0, rec[7] or "")

            def save_changes():
                new_name = name_entry.get()
                new_surname = surname_entry.get()
                new_problem_text = problem_var.get() or problem_entry_edit.get()
                new_deadline_hours = deadline_entry.get().strip()
                # Если пользователь оставил поле пустым, сохраняем исходный срок
                if not new_deadline_hours:
                    new_deadline_hours = _prefill_deadline or ""

                # Если срок указан — валидируем
                if new_deadline_hours:
                    try:
                        deadline_int = int(new_deadline_hours)
                        if deadline_int <= 0:
                            messagebox.showwarning("Ошибка", "Срок выполнения должен быть положительным числом")
                            return
                    except ValueError:
                        messagebox.showwarning("Ошибка", "Срок выполнения должен быть числом")
                        return

                # Очищаем текст проблемы от категорий в скобках (если они есть)
                new_problem_text_cleaned = clean_problem_text(new_problem_text) if new_problem_text else ""
                
                # Включаем срок в текст только если он заполнен
                if new_problem_text_cleaned:
                    new_problem = f"{new_problem_text_cleaned} (срок выполнения {new_deadline_hours} ч)" if new_deadline_hours else new_problem_text_cleaned
                else:
                    new_problem = ""
                new_phone = phone_entry.get()
                new_address = address_entry.get()
                try:
                    new_assignment = assignment_entry.get_date().strftime("%Y-%m-%d") if assignment_entry.get() else ""
                except Exception:
                    new_assignment = ""
                base_status = status_var.get()
                status_timestamp = datetime.now().strftime("%d.%m.%Y %H:%M")
                new_status = f"{base_status} ({status_timestamp})"
                # Имя и фамилия больше не обязательны
                new_improvement = improvement_entry.get()
                # Обработка номера бригады: добавляем ".бр" если номер указан
                new_brigade_number = brigade_var.get().strip()
                if new_brigade_number:
                    if not new_brigade_number.endswith(".бр"):
                        new_brigade_number = f"{new_brigade_number}.бр"
                new_category = category_var_edit.get()
                # Получаем или создаём category_id
                new_category_id = None
                if new_category:
                    try:
                        cursor.execute("INSERT OR IGNORE INTO common.categories (name) VALUES (?)", (new_category,))
                        cursor.execute("SELECT id FROM common.categories WHERE name=?", (new_category,))
                        _r = cursor.fetchone()
                        new_category_id = _r[0] if _r else None
                    except Exception:
                        pass
                new_deadline_int = int(new_deadline_hours) if new_deadline_hours and str(new_deadline_hours).isdigit() else None
                cursor.execute(
                    "UPDATE records SET name=?, surname=?, problem=?, problem_text=?, deadline_hours=?, "
                    "phone=?, address=?, assignment_date=?, status=?, status_clean=?, status_changed_at=?, "
                    "improvement=?, brigade_number=?, category=?, category_id=? WHERE id=?",
                    (new_name, new_surname, new_problem, new_problem_text_cleaned, new_deadline_int,
                     new_phone, new_address, new_assignment, new_status, base_status, status_timestamp,
                     new_improvement, new_brigade_number, new_category, new_category_id, record_id))
                conn.commit()
                refresh_records_default()
                edit_win.destroy()
                try:
                    refresh_recent()
                except Exception:
                    pass

            btns = ttk.Frame(edit_win, padding="10")
            btns.grid(row=1, column=0, sticky=(tk.E, tk.W))
            btns.columnconfigure(0, weight=1)
            ttk.Label(btns, text="").grid(row=0, column=0, sticky=tk.W)
            ttk.Button(btns, text="Сохранить", command=save_changes).grid(row=0, column=1, padx=5, sticky=tk.E)
            ttk.Button(btns, text="Отмена", command=edit_win.destroy).grid(row=0, column=2, padx=5, sticky=tk.E)

        # Функция изменения статуса - доступна всем пользователям
        def update_status_record():
            selection = records_tree.selection()
            if not selection:
                messagebox.showwarning("Ошибка", "Выберите запись")
                return
            record_id = records_tree.item(selection[0])['values'][0]

            cursor.execute("SELECT status FROM records WHERE id=?", (record_id,))
            rec = cursor.fetchone()
            try:
                curr_status = re.sub(r"\s*\(\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2}\)\s*$", "", ((rec[0] if rec else "не выполнено") or "не выполнено")).strip()
            except Exception:
                curr_status = (rec[0] if rec else "не выполнено") or "не выполнено"

            st_win = tk.Toplevel(main)
            setup_scaling()
            st_win.title("Изменить статус")
            try:
                fit_and_center_window(st_win, min_width=320, min_height=120)
            except Exception:
                pass

            frm = ttk.Frame(st_win, padding="16")
            frm.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            st_win.columnconfigure(0, weight=1)
            st_win.rowconfigure(0, weight=1)

            ttk.Label(frm, text="Статус:").grid(row=0, column=0, sticky=tk.W, pady=6)
            st_var = tk.StringVar(value=curr_status)
            st_combo = ttk.Combobox(frm, textvariable=st_var, values=["в работе", "не выполнено", "выполнено"], state="readonly", width=22)
            st_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=6)

            def save_status():
                base_status = st_var.get()
                ts = datetime.now().strftime("%d.%m.%Y %H:%M")
                new_status = f"{base_status} ({ts})"
                cursor.execute(
                    "UPDATE records SET status=?, status_clean=?, status_changed_at=? WHERE id=?",
                    (new_status, base_status, ts, record_id))
                conn.commit()
                refresh_records_default()
                st_win.destroy()
                try:
                    refresh_recent()
                except Exception:
                    pass

            btns = ttk.Frame(st_win, padding="10")
            btns.grid(row=1, column=0, sticky=(tk.E, tk.W))
            ttk.Button(btns, text="Сохранить", command=save_status).grid(row=0, column=0, padx=5)
            ttk.Button(btns, text="Отмена", command=st_win.destroy).grid(row=0, column=1, padx=5)

        # Кнопки редактирования и изменения статуса доступны всем. Удаление - только свои записи (для обычных пользователей)
        button_frame = ttk.Frame(records_window)
        button_frame.grid(row=3, column=0, pady=10)
        btn_edit = ttk.Button(button_frame, text="Редактировать", command=edit_record)
        btn_edit.pack(side="left", padx=10)
        btn_delete = ttk.Button(button_frame, text="Удалить", command=delete_record)
        btn_delete.pack(side="left", padx=10)
        btn_status = ttk.Button(button_frame, text="Изменить статус", command=update_status_record)
        btn_status.pack(side="left", padx=10)

        # Первичное заполнение — без сохранения фильтров (показываем все записи)
        try:
            rows = get_filtered_rows_for(None, None, None, None, None, None)
            populate_tree_from_rows(rows)
        except Exception:
            try:
                rows = get_filtered_rows(None)
                populate_tree_from_rows(rows)
            except Exception:
                pass

        def on_close():
            nonlocal records_window, records_tree
            try:
                records_window.destroy()
            finally:
                records_window = None
                records_tree = None

        records_window.protocol("WM_DELETE_WINDOW", on_close)

    # ---- Фоновый опрос базы и автообновление ----
    def _poll_updates():
        nonlocal last_seen_recent_sig, records_window
        try:
            current_sig = _get_recent_signature()
        except Exception:
            current_sig = None
        if current_sig and current_sig != last_seen_recent_sig:
            last_seen_recent_sig = current_sig
            try:
                refresh_recent()
            except Exception:
                pass
            try:
                if records_window and records_window.winfo_exists():
                    refresh_records_default()
            except Exception:
                pass
        try:
            main.after(3000, _poll_updates)
        except Exception:
            pass

    try:
        main.after(1000, _poll_updates)
    except Exception:
        pass

    # На главном экране оставим только кнопку открытия списка
    open_list_frame = ttk.Frame(main)
    open_list_frame.grid(row=3, column=0, columnspan=2, padx=16, pady=8, sticky="ew")
    try:
        open_list_frame.columnconfigure(0, weight=1)
    except Exception:
        pass
    ttk.Button(open_list_frame, text="Открыть список данных", command=open_records_window).grid(row=0, column=0, pady=2)
    ttk.Button(open_list_frame, text="Аналитика и прогнозирование",
               command=lambda: open_analytics_window(main)).grid(row=0, column=1, pady=2, padx=(16, 0))


    # ---- Функция отображения записей ----
    # При старте не показываем список; он открывается отдельной кнопкой
    main.mainloop()

def open_analytics_window(parent=None):
    """Окно аналитики: прогноз загрузки и проблемные адреса."""

    from collections import defaultdict

    def _collect_records():
        """Возвращает все записи из всех доступных помесячных БД."""
        base_dir = _load_db_base_dir()
        rows = []
        for db_file in sorted(Path(base_dir).glob("app_*.db")):
            if db_file.name == "app.db":
                continue
            try:
                tmp = sqlite3.connect(str(db_file))
                tc = tmp.cursor()
                ensure_schema_for_connection(tmp, tc)
                tc.execute(
                    "SELECT id, address, category, problem, "
                    "COALESCE(created_at, date) as dt, status FROM records"
                )
                rows.extend(tc.fetchall())
                tmp.close()
            except Exception:
                pass
        return rows

    def _linreg(x_vals, y_vals):
        """Простая линейная регрессия. Возвращает (slope, intercept)."""
        n = len(x_vals)
        if n < 2:
            return 0, (sum(y_vals) / n if y_vals else 0)
        xm = sum(x_vals) / n
        ym = sum(y_vals) / n
        num = sum((x_vals[i] - xm) * (y_vals[i] - ym) for i in range(n))
        den = sum((x_vals[i] - xm) ** 2 for i in range(n))
        if den == 0:
            return 0, ym
        slope = num / den
        return slope, ym - slope * xm

    def _draw_bars(canvas, labels, values, forecast_from=None):
        """Рисует вертикальную столбчатую диаграмму на Canvas."""
        canvas.delete('all')
        canvas.update_idletasks()
        W = canvas.winfo_width() or 860
        H = canvas.winfo_height() or 300

        if not values:
            canvas.create_text(W // 2, H // 2, text="Нет данных",
                               font=('Segoe UI', 12), fill='#888')
            return

        ML, MR, MT, MB = 46, 20, 30, 64
        cw = W - ML - MR
        ch = H - MT - MB
        n = len(values)
        max_v = max(values) or 1
        bar_w = max(4, cw // n - 3)
        fact_c, fore_c = '#1e88e5', '#ef6c00'

        # Сетка и ось Y
        canvas.create_line(ML, MT, ML, H - MB, fill='#444', width=2)
        canvas.create_line(ML, H - MB, W - MR, H - MB, fill='#444', width=2)
        for i in range(5):
            yv = max_v * i / 4
            yp = H - MB - int(ch * i / 4)
            canvas.create_line(ML - 4, yp, ML, yp, fill='#888')
            canvas.create_text(ML - 6, yp, text=str(int(yv)),
                               anchor='e', font=('Segoe UI', 8), fill='#555')
            canvas.create_line(ML, yp, W - MR, yp, fill='#eee', dash=(2, 4))

        # Столбцы
        step = cw // n
        for i, (lbl, v) in enumerate(zip(labels, values)):
            x1 = ML + i * step + 2
            x2 = x1 + bar_w
            bh = int(ch * v / max_v) if max_v else 0
            y1 = H - MB - bh
            y2 = H - MB
            color = fore_c if (forecast_from is not None and i >= forecast_from) else fact_c
            canvas.create_rectangle(x1, y1, x2, y2, fill=color, outline='')
            if v > 0:
                canvas.create_text((x1 + x2) // 2, y1 - 3, text=str(int(v)),
                                   font=('Segoe UI', 8), fill='#222', anchor='s')
            # Подпись по оси X (каждая вторая при тесноте)
            if n <= 24 or i % 2 == 0:
                canvas.create_text((x1 + x2) // 2, H - MB + 4, text=lbl,
                                   font=('Segoe UI', 7), fill='#555', anchor='n')

        # Легенда
        if forecast_from is not None:
            lx = W - 160
            canvas.create_rectangle(lx, 8, lx + 14, 20, fill=fact_c, outline='')
            canvas.create_text(lx + 18, 14, text="Факт", anchor='w',
                               font=('Segoe UI', 8), fill='#333')
            canvas.create_rectangle(lx + 60, 8, lx + 74, 20, fill=fore_c, outline='')
            canvas.create_text(lx + 78, 14, text="Прогноз", anchor='w',
                               font=('Segoe UI', 8), fill='#333')

    def _draw_hbars(canvas, labels, values):
        """Рисует горизонтальную столбчатую диаграмму на Canvas."""
        canvas.delete('all')
        canvas.update_idletasks()
        W = canvas.winfo_width() or 860
        H = canvas.winfo_height() or 400

        if not values:
            canvas.create_text(W // 2, H // 2, text="Нет данных",
                               font=('Segoe UI', 12), fill='#888')
            return

        n = len(values)
        ML, MR, MT, MB = 240, 70, 14, 14
        ch = H - MT - MB
        cw = W - ML - MR
        max_v = max(values) or 1
        bar_h = max(10, ch // n - 4)
        palette = ['#1e88e5', '#43a047', '#fb8c00', '#e53935', '#8e24aa',
                   '#00897b', '#3949ab', '#f06292', '#0097a7', '#6d4c41']

        canvas.create_line(ML, MT, ML, H - MB, fill='#444', width=2)

        step = ch // n
        for i, (lbl, v) in enumerate(zip(labels, values)):
            yc = MT + i * step + step // 2
            bw = int(cw * v / max_v) if max_v else 0
            color = palette[i % len(palette)]
            canvas.create_rectangle(ML, yc - bar_h // 2,
                                    ML + bw, yc + bar_h // 2,
                                    fill=color, outline='')
            short = lbl[:36] + "…" if len(lbl) > 36 else lbl
            canvas.create_text(ML - 4, yc, text=short,
                               anchor='e', font=('Segoe UI', 8), fill='#222')
            canvas.create_text(ML + bw + 5, yc, text=str(int(v)),
                               anchor='w', font=('Segoe UI', 8, 'bold'), fill='#222')

    # ── Окно ──────────────────────────────────────────────────────────────
    win = tk.Toplevel(parent) if parent else tk.Toplevel()
    win.title("Аналитика и прогнозирование")
    setup_scaling()
    try:
        win.state('zoomed')
    except Exception:
        pass

    notebook = ttk.Notebook(win)
    notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=8)

    # ── Вкладка 1: Прогноз загрузки ───────────────────────────────────────
    tab1 = ttk.Frame(notebook)
    notebook.add(tab1, text="  Прогноз загрузки  ")
    tab1.rowconfigure(1, weight=1)
    tab1.columnconfigure(0, weight=1)

    ctrl1 = ttk.Frame(tab1, padding=(8, 6))
    ctrl1.grid(row=0, column=0, sticky='ew')

    ttk.Label(ctrl1, text="С:").pack(side=tk.LEFT, padx=(0, 4))
    date_from_entry = DateEntry(ctrl1, date_pattern='dd.mm.yyyy', width=10)
    date_from_entry.set_date(datetime.now() - timedelta(days=180))
    date_from_entry.pack(side=tk.LEFT)

    ttk.Label(ctrl1, text="По:").pack(side=tk.LEFT, padx=(8, 4))
    date_to_entry = DateEntry(ctrl1, date_pattern='dd.mm.yyyy', width=10)
    date_to_entry.set_date(datetime.now())
    date_to_entry.pack(side=tk.LEFT, padx=(0, 16))

    ttk.Label(ctrl1, text="Группировка:").pack(side=tk.LEFT, padx=(0, 4))
    period_var = tk.StringVar(value="week")
    ttk.Radiobutton(ctrl1, text="По неделям", variable=period_var,
                    value="week").pack(side=tk.LEFT)
    ttk.Radiobutton(ctrl1, text="По дням", variable=period_var,
                    value="day").pack(side=tk.LEFT, padx=(4, 16))

    ttk.Label(ctrl1, text="Прогноз на:").pack(side=tk.LEFT, padx=(0, 4))
    fperiods_var = tk.IntVar(value=4)
    ttk.Spinbox(ctrl1, from_=1, to=16, textvariable=fperiods_var,
                width=4).pack(side=tk.LEFT)
    ttk.Label(ctrl1, text="периодов вперёд").pack(side=tk.LEFT, padx=(4, 0))
    ttk.Button(ctrl1, text="Применить",
               command=lambda: _refresh_forecast()).pack(side=tk.LEFT, padx=(16, 0))

    desc_frame = ttk.Frame(tab1)
    desc_frame.grid(row=1, column=0, sticky='ew', padx=12, pady=(4, 0))
    ttk.Label(desc_frame,
              text="Прогноз загрузки показывает количество заявок за каждую неделю выбранного периода (синие столбцы)\n"
                   "и прогнозируемое количество заявок на последующие недели на основе линейной регрессии (оранжевые столбцы).",
              font=('Segoe UI', 9), foreground='#555').pack(anchor='w')

    chart_frame1 = ttk.LabelFrame(tab1, text="График", padding=4)
    chart_frame1.grid(row=2, column=0, sticky='nsew', padx=8, pady=4)
    tab1.rowconfigure(2, weight=1)
    chart_frame1.rowconfigure(0, weight=1)
    chart_frame1.columnconfigure(0, weight=1)
    chart_canvas1 = tk.Canvas(chart_frame1, bg='white')
    chart_canvas1.grid(row=0, column=0, sticky='nsew')

    # ── Вкладка 2: Проблемные адреса ─────────────────────────────────────
    tab2 = ttk.Frame(notebook)
    notebook.add(tab2, text="  Проблемные адреса  ")
    tab2.rowconfigure(1, weight=1)
    tab2.columnconfigure(0, weight=1)

    ctrl2 = ttk.Frame(tab2, padding=(8, 6))
    ctrl2.grid(row=0, column=0, sticky='ew')
    ttk.Label(ctrl2, text="Показать топ:").pack(side=tk.LEFT, padx=(0, 4))
    topn_var = tk.IntVar(value=10)
    ttk.Spinbox(ctrl2, from_=5, to=50, textvariable=topn_var,
                width=4).pack(side=tk.LEFT)
    ttk.Label(ctrl2, text="адресов").pack(side=tk.LEFT, padx=(4, 0))

    paned2 = ttk.PanedWindow(tab2, orient=tk.VERTICAL)
    paned2.grid(row=1, column=0, sticky='nsew', padx=8, pady=4)

    chart_frame2 = ttk.LabelFrame(paned2, text="Рейтинг адресов", padding=4)
    paned2.add(chart_frame2, weight=3)
    chart_canvas2 = tk.Canvas(chart_frame2, bg='white')
    chart_canvas2.pack(fill=tk.BOTH, expand=True)

    table_frame2 = ttk.LabelFrame(paned2, text="Данные", padding=4)
    paned2.add(table_frame2, weight=2)
    cols2 = ("Место", "Адрес", "Заявок", "Категории")
    tree2 = ttk.Treeview(table_frame2, columns=cols2, show='headings', height=7)
    tree2.heading("Место", text="#")
    tree2.column("Место", width=44, anchor='center')
    tree2.heading("Адрес", text="Адрес")
    tree2.column("Адрес", width=400)
    tree2.heading("Заявок", text="Заявок")
    tree2.column("Заявок", width=80, anchor='center')
    tree2.heading("Категории", text="Категории")
    tree2.column("Категории", width=260)
    sb2y = ttk.Scrollbar(table_frame2, orient='vertical', command=tree2.yview)
    tree2.configure(yscrollcommand=sb2y.set)
    tree2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    sb2y.pack(side=tk.RIGHT, fill=tk.Y)

    # ── Кнопки ────────────────────────────────────────────────────────────
    btn_bar = ttk.Frame(win)
    btn_bar.pack(fill=tk.X, padx=10, pady=(0, 8))
    ttk.Button(btn_bar, text="Обновить", command=lambda: _refresh_all()).pack(
        side=tk.LEFT, padx=4)
    ttk.Button(btn_bar, text="Закрыть", command=win.destroy).pack(
        side=tk.LEFT, padx=4)

    # ── Логика обновления ─────────────────────────────────────────────────
    _cached_records = [None]   # список из одного элемента — кэш записей

    def _get_records():
        if _cached_records[0] is None:
            _cached_records[0] = _collect_records()
        return _cached_records[0]

    def _refresh_forecast():
        records = _get_records()
        period = period_var.get()
        n_fore = fperiods_var.get()

        # Фильтр по выбранному диапазону дат
        try:
            dt_from = date_from_entry.get_date()
            dt_to = date_to_entry.get_date()
        except Exception:
            dt_from = dt_to = None

        counts = defaultdict(int)
        for row in records:
            dt_str = row[4]
            if not dt_str:
                continue
            try:
                dt = (datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
                      if ' ' in dt_str
                      else datetime.strptime(dt_str, "%Y-%m-%d"))
                if dt_from and dt.date() < dt_from:
                    continue
                if dt_to and dt.date() > dt_to:
                    continue
                key = dt.strftime("%Y-W%W") if period == "week" else dt.strftime("%Y-%m-%d")
            except Exception:
                continue
            counts[key] += 1

        sorted_keys = sorted(counts)
        y_vals = [counts[k] for k in sorted_keys]
        x_vals = list(range(len(sorted_keys)))

        slope, intercept = _linreg(x_vals, y_vals)

        # Прогнозные периоды
        fore_labels, fore_vals = [], []
        base = len(sorted_keys)
        if period == "week" and sorted_keys:
            try:
                yr, wk = sorted_keys[-1].split('-W')
                last_dt = datetime.strptime(f"{yr}-W{int(wk):02d}-1", "%Y-W%W-%w")
                for i in range(1, n_fore + 1):
                    nd = last_dt + timedelta(weeks=i)
                    fore_labels.append(nd.strftime("%Y-W%W"))
                    fore_vals.append(max(0, round(slope * (base + i - 1) + intercept)))
            except Exception:
                for i in range(1, n_fore + 1):
                    fore_labels.append(f"+{i} нед.")
                    fore_vals.append(max(0, round(slope * (base + i - 1) + intercept)))
        else:
            for i in range(1, n_fore + 1):
                fore_labels.append(f"Прогноз +{i}")
                fore_vals.append(max(0, round(slope * (base + i - 1) + intercept)))

        all_labels = sorted_keys + fore_labels
        all_vals = y_vals + fore_vals
        fc_start = len(sorted_keys)

        def _short(lbl):
            if '-W' in lbl:
                try:
                    yr, wk = lbl.split('-W')
                    dt_start = datetime.strptime(f"{yr}-W{int(wk):02d}-1", "%Y-W%W-%w")
                    dt_end = dt_start + timedelta(days=6)
                    return f"{dt_start.strftime('%d.%m')}-{dt_end.strftime('%d.%m')}"
                except Exception:
                    return lbl
            if lbl.count('-') == 2:
                parts = lbl.split('-')
                return f"{parts[2]}.{parts[1]}"
            return lbl

        short_labels = [_short(l) for l in all_labels]
        _draw_bars(chart_canvas1, short_labels, all_vals, fc_start)

    def _normalize_addr(raw):
        """Нормализует адрес для группировки: нижний регистр, единые сокращения, лишние пробелы/знаки."""
        if not raw:
            return ""
        s = raw.strip().lower()
        # Убираем лишние пробелы вокруг знаков препинания и внутри строки
        s = re.sub(r'\s+', ' ', s)
        s = re.sub(r'\s*[,;]\s*', ', ', s)
        s = re.sub(r'\s*\.\s*', '.', s)
        # Нормализуем типовые сокращения улиц
        abbr = [
            (r'\bулица\b',       'ул.'),
            (r'\bул\b\.?',       'ул.'),
            (r'\bпроспект\b',    'пр.'),
            (r'\bпр-т\b\.?',     'пр.'),
            (r'\bпр\b\.?',       'пр.'),
            (r'\bпереулок\b',    'пер.'),
            (r'\bпер\b\.?',      'пер.'),
            (r'\bбульвар\b',     'бул.'),
            (r'\bбул\b\.?',      'бул.'),
            (r'\bшоссе\b',       'ш.'),
            (r'\bш\b\.?',        'ш.'),
            (r'\bдом\b',         'д.'),
            (r'\bд\b\.?(?=\s*\d)', 'д.'),
            (r'\bкорпус\b',      'корп.'),
            (r'\bкорп\b\.?',     'корп.'),
            (r'\bквартира\b',    'кв.'),
            (r'\bкв\b\.?',       'кв.'),
        ]
        for pattern, repl in abbr:
            s = re.sub(pattern, repl, s)
        # Убираем двойные точки и финальный мусор
        s = re.sub(r'\.{2,}', '.', s)
        s = re.sub(r'\s+', ' ', s).strip().strip(',').strip()
        return s

    def _refresh_addresses():
        records = _get_records()
        top_n = topn_var.get()

        addr_cnt = defaultdict(int)
        addr_cats = defaultdict(set)
        # Для отображения сохраняем наиболее часто встречающийся оригинальный вариант
        addr_display = defaultdict(lambda: defaultdict(int))

        for row in records:
            raw = (row[1] or "").strip()
            cat = (row[2] or "").strip()
            if not raw:
                continue
            key = _normalize_addr(raw)
            if not key:
                continue
            addr_cnt[key] += 1
            addr_display[key][raw] += 1
            if cat:
                addr_cats[key].add(cat)

        top = sorted(addr_cnt.items(), key=lambda x: x[1], reverse=True)[:top_n]

        # Берём самый частый оригинальный вариант адреса для отображения
        display_labels = []
        for key, _ in top:
            best = max(addr_display[key], key=addr_display[key].get)
            display_labels.append(best)

        values = [a[1] for a in top]
        _draw_hbars(chart_canvas2, display_labels, values)

        for item in tree2.get_children():
            tree2.delete(item)
        for rank, ((key, cnt), label) in enumerate(zip(top, display_labels), 1):
            cats = ", ".join(sorted(addr_cats[key])) or "—"
            tree2.insert('', 'end', values=(rank, label, cnt, cats))

    def _refresh_all():
        _cached_records[0] = None  # сбрасываем кэш
        try:
            _refresh_forecast()
        except Exception:
            pass
        try:
            _refresh_addresses()
        except Exception:
            pass

    # Перерисовка при изменении размера окна
    def _on_resize_forecast(event):
        try:
            _refresh_forecast()
        except Exception:
            pass

    def _on_resize_addresses(event):
        try:
            _refresh_addresses()
        except Exception:
            pass

    chart_canvas1.bind('<Configure>', _on_resize_forecast)
    chart_canvas2.bind('<Configure>', _on_resize_addresses)

    # Обновляем при переключении вкладок
    def _on_tab_change(event):
        idx = notebook.index("current")
        win.after(80, _refresh_forecast if idx == 0 else _refresh_addresses)

    notebook.bind("<<NotebookTabChanged>>", _on_tab_change)

    # Первоначальная загрузка
    win.after(300, _refresh_all)


def open_login_window():
    global login_window
    # ===== Окно входа =====
    login_window = tk.Tk()
    login_window.title("Вход в систему")
    setup_scaling()
    try:
        _login_icon = load_resized_image("logo2.png", target_height=32)
        if _login_icon is not None:
            login_window.iconphoto(True, _login_icon)
            login_window._icon_img = _login_icon  # удерживаем ссылку
    except Exception:
        pass

    # Настройка масштабирования для окна входа
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
        ctypes.windll.user32.SetProcessDPIAware()
    except:
        pass

    style = ttk.Style()
    try:
        style.theme_use('clam')
    except:
        pass
    style.configure('TButton', padding=8, font=('Segoe UI', 10))
    style.configure('TLabel', font=('Segoe UI', 10))
    style.configure('TEntry', padding=6, font=('Segoe UI', 10))

    # Настройка размеров окна и центрирование
    window_width = 420
    window_height = 240
    screen_width = login_window.winfo_screenwidth()
    screen_height = login_window.winfo_screenheight()
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)
    login_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
    login_window.resizable(False, False)

    # Настройка весов
    login_window.columnconfigure(0, weight=1)
    login_window.rowconfigure(0, weight=1)

    main_frame = ttk.Frame(login_window, padding="30")
    main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
    main_frame.columnconfigure(1, weight=1)

    ttk.Label(main_frame, text="Логин:", font=('Segoe UI', 10)).grid(row=0, column=0, sticky=tk.W, pady=10)
    global entry_login
    entry_login = ttk.Entry(main_frame, width=25, font=('Segoe UI', 10))
    entry_login.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(15, 0), pady=10)

    ttk.Label(main_frame, text="Пароль:", font=('Segoe UI', 10)).grid(row=1, column=0, sticky=tk.W, pady=10)
    global entry_password
    entry_password = ttk.Entry(main_frame, show="*", width=25, font=('Segoe UI', 10))
    entry_password.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(15, 0), pady=10)
    # Иконка-глаз показать/скрыть пароль (кликаемая метка)
    def toggle_password():
        try:
            if entry_password.cget('show') == '*':
                entry_password.config(show='')
                eye_toggle_label.config(text='😌')
            else:
                entry_password.config(show='*')
                eye_toggle_label.config(text='👁')
        except Exception:
            pass
    eye_toggle_label = tk.Label(main_frame, text="👁", font=('Segoe UI', 10), width=2, anchor='w')
    try:
        eye_toggle_label.configure(cursor="hand2")
    except Exception:
        pass
    eye_toggle_label.grid(row=1, column=2, sticky=tk.W, padx=(8, 0))
    eye_toggle_label.bind("<Button-1>", lambda e: toggle_password())

    # Чекбокс «Запомнить пользователя» и подстановка логина
    global remember_var
    remember_var = tk.BooleanVar(value=False)
    try:
        if os.path.exists(".remember_user"):
            with open(".remember_user", "r", encoding="utf-8") as f:
                lines = f.read().splitlines()
            saved_user = lines[0] if len(lines) > 0 else ""
            saved_pass = lines[1] if len(lines) > 1 else ""
            if saved_user:
                entry_login.delete(0, tk.END)
                entry_login.insert(0, saved_user)
                remember_var.set(True)
            entry_password.delete(0, tk.END)
            entry_password.insert(0, saved_pass or "")
    except Exception:
        pass

    try:
        chk = tk.Checkbutton(main_frame, text="Запомнить пользователя и пароль", variable=remember_var, onvalue=True, offvalue=False)
        try:
            chk.configure(font=('Segoe UI', 10))
        except Exception:
            pass
        chk.grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=(2, 0))
    except Exception:
        ttk.Checkbutton(main_frame, text="Запомнить пользователя и пароль", variable=remember_var).grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=(2, 0))

    ttk.Button(main_frame, text="Войти", command=login, style='Accent.TButton').grid(row=3, column=0, columnspan=2, pady=16)

    # Привязка Enter к кнопке входа
    login_window.bind('<Return>', lambda event: login())

    login_window.mainloop()

open_login_window()
